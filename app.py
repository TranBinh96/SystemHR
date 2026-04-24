from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_admin import Admin
from flask_admin.contrib.sqla import ModelView
from flask_migrate import Migrate
from flask_jwt_extended import JWTManager
from flask_wtf.csrf import CSRFProtect
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger

# Import timezone utilities
from utils.timezone_utils import now, today, utcnow, format_local_datetime, format_local_date

from models import db, User, OvertimeRequest, LeaveRequest, MealRegistration, ExitEntryRequest
from config import Config
from translations import get_translation
from forms import LoginForm, RegisterForm, OvertimeForm, ChangePasswordForm

# Initialize Flask app
app = Flask(__name__)
app.config.from_object(Config)

# Disable template caching for development
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

# Initialize extensions
db.init_app(app)
migrate = Migrate(app, db)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'

# Initialize CSRF Protection
# csrf = CSRFProtect(app)  # Tắt tạm thời để debug

# Initialize JWT
jwt = JWTManager(app)

# Register API Blueprint
from api import api_bp
app.register_blueprint(api_bp)

# Initialize Flask-Admin
admin = Admin(app, name='OKI HR Admin', template_mode='bootstrap4')
admin.add_view(ModelView(User, db.session))
admin.add_view(ModelView(OvertimeRequest, db.session))
admin.add_view(ModelView(MealRegistration, db.session))

# Auto-create database and tables on first run
def init_database():
    """
    Initialize database and tables if they don't exist
    Automatically detects and creates new tables when models are added
    """
    try:
        with app.app_context():
            # Get existing tables before creation
            inspector = db.inspect(db.engine)
            existing_tables = inspector.get_table_names()
            
            # Create all tables (SQLAlchemy automatically detects all models)
            db.create_all()
            
            # Get tables after creation to see what's new
            inspector = db.inspect(db.engine)
            current_tables = inspector.get_table_names()
            new_tables = set(current_tables) - set(existing_tables)
            
            if new_tables:
                print(f"🆕 New tables created: {', '.join(new_tables)}")
            
            # Check if we need to create default users
            if User.query.count() == 0:
                print("Creating default users...")
                admin = User(
                    employee_id='ADMIN',
                    name='Administrator',
                    email='admin@okivietnam.com',
                    password=generate_password_hash('admin123'),
                    department='IT',
                    role='admin'
                )
                sample_user = User(
                    employee_id='EMP001',
                    name='Nguyen Van A',
                    email='nguyenvana@okivietnam.com',
                    password=generate_password_hash('password123'),
                    department='production',
                    role='user'
                )
                db.session.add(admin)
                db.session.add(sample_user)
                db.session.commit()
                print("✓ Default users created")
    except Exception as e:
        print(f"Database initialization: {e}")

# Initialize database on startup
init_database()

# Print configuration on startup (for debugging)
if app.config['DEBUG']:
    Config.print_config()

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Auto logout middleware - check last activity
@app.before_request
def check_user_activity():
    """Check if user should be logged out due to inactivity"""
    if current_user.is_authenticated:
        # Skip for static files, API endpoints, and logout
        if request.endpoint and (
            request.endpoint.startswith('static') or 
            request.endpoint.startswith('api') or 
            request.endpoint == 'logout'
        ):
            return
        
        # Check last activity
        if current_user.last_activity:
            inactive_duration = utcnow() - current_user.last_activity  # Use timezone-aware UTC
            max_inactive = timedelta(seconds=app.config['AUTO_LOGOUT_DURATION'])
            
            if inactive_duration > max_inactive:
                logout_user()
                session.clear()
                flash('Bạn đã bị đăng xuất do không hoạt động trong 3 tuần', 'warning')
                return redirect(url_for('login'))
        
        # Update last activity
        try:
            current_user.last_activity = utcnow()  # Use timezone-aware UTC
            db.session.commit()
        except:
            db.session.rollback()

# Language middleware
@app.before_request
def set_language():
    if 'lang' not in session:
        session['lang'] = 'vi'  # Default language

@app.context_processor
def inject_translations():
    """Inject translations into all templates"""
    lang = session.get('lang', 'vi')
    
    # Check if user is a manager (has subordinates based on level)
    is_manager = False
    if current_user.is_authenticated:
        # New level-based system: user is manager if they have subordinates
        is_manager = current_user.is_manager() or current_user.role == 'admin'
    
    return dict(t=get_translation(lang), current_lang=lang, is_manager=is_manager)

# Routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/set-language/<lang>')
def set_language_route(lang):
    """Thay đổi ngôn ngữ"""
    if lang in ['vi', 'en', 'ja']:
        session['lang'] = lang
    return redirect(request.referrer or url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(employee_id=form.employee_id.data).first()
        if user and check_password_hash(user.password, form.password.data):
            # Auto-activate if work_status is 'working'
            if user.work_status == 'working' and not user.is_active:
                user.is_active = True
                db.session.commit()
            
            # Check if account is active
            if not user.is_active:
                flash('Tài khoản của bạn đã bị vô hiệu hóa. Vui lòng liên hệ quản trị viên.', 'error')
                return render_template('login.html', form=form)
            
            # Update last activity on login
            user.last_activity = utcnow()  # Use timezone-aware UTC
            db.session.commit()
            
            # Login user with remember me for 3 weeks
            login_user(user, remember=True, duration=timedelta(weeks=3))
            session['user_name'] = user.name
            session.permanent = True  # Make session permanent (3 weeks)
            
            next_page = request.args.get('next')
            return redirect(next_page or url_for('dashboard'))
        else:
            flash('Invalid employee ID or password', 'error')
    
    return render_template('login.html', form=form)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = RegisterForm()
    if form.validate_on_submit():
        hashed_password = generate_password_hash(form.password.data)
        user = User(
            employee_id=form.employee_id.data,
            name=form.name.data,
            email=form.email.data,
            department=form.department.data,
            password=hashed_password
        )
        db.session.add(user)
        db.session.commit()
        flash('Registration successful! Please login.', 'success')
        return redirect(url_for('login'))
    
    return render_template('register.html', form=form)

@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        identifier = request.form.get('identifier')
        
        if not identifier:
            flash('Please provide your email or employee ID', 'error')
            return render_template('forgot_password.html')
        
        # TODO: Send reset email
        flash('Password reset instructions have been sent to your email', 'success')
        return redirect(url_for('login'))
    
    return render_template('forgot_password.html')

@app.route('/dashboard')
@login_required
def dashboard():
    # Check if admin
    if current_user.role == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Check if user can approve (has can_approve permission)
    is_manager = current_user.can_approve
    
    return render_template('dashboard.html', user=current_user.name, is_manager=is_manager)

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    # Only admin can access
    if current_user.role != 'admin':
        flash('You do not have permission to access this page', 'error')
        return render_template('dashboard.html', user=current_user.name)
    
    # Calculate stats - only count users with work_status='working' and employee_id starting with 10 or 20
    from sqlalchemy import or_, and_
    
    total_users = User.query.filter(
        and_(
            User.work_status == 'working',
            or_(
                User.employee_id.like('10%'),
                User.employee_id.like('20%')
            )
        )
    ).count()
    
    working_users = User.query.filter(
        and_(
            User.work_status == 'working',
            or_(
                User.employee_id.like('10%'),
                User.employee_id.like('20%')
            )
        )
    ).count()
    
    business_trip_users = User.query.filter(
        and_(
            User.work_status == 'business_trip',
            or_(
                User.employee_id.like('10%'),
                User.employee_id.like('20%')
            )
        )
    ).count()
    
    # Calculate overtime stats for today
    from datetime import date
    today_date = today()  # Use timezone-aware today
    
    today_overtime = OvertimeRequest.query.filter(
        OvertimeRequest.overtime_date == today_date,
        OvertimeRequest.status == 'approved'
    ).all()
    
    overtime_count = len(today_overtime)
    overtime_hours = sum(float(ot.total_hours) for ot in today_overtime)
    
    # Calculate meal stats for tomorrow
    from datetime import timedelta
    from models import Menu
    tomorrow = today_date + timedelta(days=1)
    
    # Get all meal registrations for tomorrow - only active users with employee_id starting with 10 or 20
    tomorrow_meals = MealRegistration.query.join(
        User,
        MealRegistration.user_id == User.id
    ).filter(
        MealRegistration.date == tomorrow,
        MealRegistration.has_meal == True,
        User.work_status == 'working',  # Chỉ đếm user đang hoạt động
        or_(
            User.employee_id.like('10%'),
            User.employee_id.like('20%')
        )
    ).all()
    
    # Count by meal type (normal vs special/improved)
    total_meals = len(tomorrow_meals)
    special_meals = 0
    normal_meals = 0
    
    for meal in tomorrow_meals:
        if meal.menu and meal.menu.is_special:
            special_meals += 1
        else:
            normal_meals += 1
    
    # Debug statements removed to prevent OSError
    
    return render_template('admin_dashboard.html', 
                         user=current_user.name,
                         total_users=total_users,
                         working_users=working_users,
                         business_trip_users=business_trip_users,
                         overtime_count=overtime_count,
                         overtime_hours=overtime_hours,
                         total_meals=total_meals,
                         normal_meals=normal_meals,
                         special_meals=special_meals)

@app.route('/admin/users')
@login_required
def admin_users():
    """User management page - admin only"""
    if current_user.role != 'admin':
        flash('Bạn không có quyền truy cập trang này', 'error')
        return redirect(url_for('dashboard'))
    
    from models import Department
    
    # Get all users
    users = User.query.all()
    
    # Sort by employee_id numerically
    # Numeric IDs first (sorted as integers), then alphabetic IDs
    def sort_key(user):
        emp_id = user.employee_id
        if emp_id.isdigit():
            return (0, int(emp_id))  # Numeric IDs: sort as integers
        else:
            return (1, emp_id.lower())  # Non-numeric IDs: sort alphabetically
    
    users.sort(key=sort_key)
    
    departments = Department.query.order_by(Department.name.asc()).all()
    
    return render_template('admin_users.html', users=users, departments=departments)

@app.route('/admin/users/list', methods=['GET'])
@login_required
def admin_users_list():
    """API endpoint to get users list as JSON"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    # Get all users
    users = User.query.all()
    
    # Sort by employee_id numerically
    def sort_key(user):
        emp_id = user.employee_id
        if emp_id.isdigit():
            return (0, int(emp_id))
        else:
            return (1, emp_id.lower())
    
    users.sort(key=sort_key)
    
    users_data = []
    for user in users:
        users_data.append({
            'id': user.id,
            'employee_id': user.employee_id,
            'name': user.name,
            'department': user.dept.name if user.dept else None,
            'role': user.role,
            'is_active': user.is_active,
            'work_status': user.work_status,
            'avatar_url': user.avatar_url,
            'can_approve': user.can_approve,
            'can_register': user.can_register
        })
    
    return {'success': True, 'users': users_data}

@app.route('/admin/approvals')
@login_required
def admin_approvals():
    """Request approval page - admin only"""
    if current_user.role != 'admin':
        flash('Bạn không có quyền truy cập trang này', 'error')
        return redirect(url_for('dashboard'))
    
    return render_template('admin_approvals.html')

@app.route('/admin/users/<int:user_id>/toggle-active', methods=['POST'])
@login_required
def toggle_user_active(user_id):
    """Toggle user active status"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    user = User.query.get_or_404(user_id)
    
    # Prevent admin from deactivating themselves
    if user.id == current_user.id:
        return {'success': False, 'message': 'Không thể vô hiệu hóa tài khoản của chính mình'}, 400
    
    user.is_active = not user.is_active
    db.session.commit()
    
    status = 'kích hoạt' if user.is_active else 'vô hiệu hóa'
    return {'success': True, 'message': f'Đã {status} user {user.name}', 'is_active': user.is_active}

@app.route('/admin/users/<int:user_id>/change-role', methods=['POST'])
@login_required
def change_user_role(user_id):
    """Change user role"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    user = User.query.get_or_404(user_id)
    
    # Prevent admin from changing their own role
    if user.id == current_user.id:
        return {'success': False, 'message': 'Không thể thay đổi quyền của chính mình'}, 400
    
    new_role = request.json.get('role')
    if new_role not in ['admin', 'user']:
        return {'success': False, 'message': 'Quyền không hợp lệ'}, 400
    
    user.role = new_role
    db.session.commit()
    
    return {'success': True, 'message': f'Đã thay đổi quyền của {user.name} thành {new_role}', 'role': user.role}

@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
def delete_user(user_id):
    """Delete user"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    user = User.query.get_or_404(user_id)
    
    # Prevent admin from deleting themselves
    if user.id == current_user.id:
        return {'success': False, 'message': 'Không thể xóa tài khoản của chính mình'}, 400
    
    user_name = user.name
    db.session.delete(user)
    db.session.commit()
    
    return {'success': True, 'message': f'Đã xóa user {user_name}'}

@app.route('/admin/users/<int:user_id>/edit', methods=['POST'])
@login_required
def edit_user(user_id):
    """Edit user information"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    from werkzeug.utils import secure_filename
    import os
    from datetime import datetime
    
    user = User.query.get_or_404(user_id)
    print(f"\n[EDIT USER] Bắt đầu edit user ID: {user_id}, Name: {user.name}")
    
    # Handle both JSON and FormData
    try:
        if request.is_json:
            data = request.json
        else:
            data = request.form.to_dict()
        
        print(f"[EDIT USER] Data received: {data}")
        
        if 'name' in data:
            user.name = data['name']
            print(f"[EDIT USER] Updated name: {user.name}")
            
        if 'employee_id' in data:
            # Check if employee_id already exists
            existing = User.query.filter(User.employee_id == data['employee_id'], User.id != user_id).first()
            if existing:
                return {'success': False, 'message': 'Mã nhân viên đã tồn tại'}, 400
            user.employee_id = data['employee_id']
            print(f"[EDIT USER] Updated employee_id: {user.employee_id}")
        
        # Handle department - find department_id from name
        if 'department' in data:
            from models import Department
            dept = Department.query.filter_by(name=data['department']).first()
            if dept:
                user.department_id = dept.id
                print(f"[EDIT USER] Updated department: {dept.name}")
            else:
                user.department_id = None
                print(f"[EDIT USER] Department not found, set to None")
        
        if 'role' in data:
            user.role = data['role']
            print(f"[EDIT USER] Updated role: {user.role}")
        
        deleted_count = 0  # Số đăng ký cơm đã xóa
        
        if 'work_status' in data:
            old_status = user.work_status
            user.work_status = data['work_status']
            print(f"[EDIT USER] Updated work_status: {old_status} → {user.work_status}")
            
            # Auto-activate khi chuyển sang 'working'
            if data['work_status'] == 'working' and not user.is_active:
                user.is_active = True
                print(f"[EDIT USER] Auto-activated is_active = True (work_status = working)")
            
            # Khi nghỉ việc: Xóa tất cả đăng ký cơm trong tương lai và vô hiệu hóa tài khoản
            if data['work_status'] == 'resigned':
                print(f"[EDIT USER] User resigned, deleting future meal registrations...")
                user.is_active = False  # Tự động vô hiệu hóa khi nghỉ việc
                print(f"[EDIT USER] Auto-deactivated is_active = False (work_status = resigned)")
                
                from models import MealRegistration
                from datetime import date
                today_date = today()
                future_registrations = MealRegistration.query.filter(
                    MealRegistration.user_id == user.id,
                    MealRegistration.date >= today_date
                ).all()
                
                deleted_count = len(future_registrations)
                for reg in future_registrations:
                    db.session.delete(reg)
                
                print(f"[EDIT USER] Deleted {deleted_count} future meal registrations")
        
        if 'is_active' in data:
            old_active = user.is_active
            # Convert string 'true'/'false' to boolean
            user.is_active = data['is_active'] in ['true', 'True', True, 1, '1']
            print(f"[EDIT USER] Updated is_active: {old_active} → {user.is_active}")
        
        # Handle permission fields
        if 'can_approve' in data:
            user.can_approve = data['can_approve'] in ['true', 'True', True, 1, '1']
            # Người có quyền phê duyệt tự động có quyền đăng ký
            if user.can_approve:
                user.can_register = True
        if 'can_register' in data:
            user.can_register = data['can_register'] in ['true', 'True', True, 1, '1']
        
        # Handle password update
        if 'password' in data and data['password']:
            from werkzeug.security import generate_password_hash
            user.password = generate_password_hash(data['password'])
            print(f"[EDIT USER] Password updated")
        
        # Handle avatar upload
        if 'avatar' in request.files:
            avatar_file = request.files['avatar']
            if avatar_file and avatar_file.filename:
                # Validate file extension
                allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
                filename = secure_filename(avatar_file.filename)
                file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
                
                if file_ext in allowed_extensions:
                    # Create upload folder if not exists
                    upload_folder = os.path.join('static', 'uploads', 'avatars')
                    os.makedirs(upload_folder, exist_ok=True)
                    
                    # Generate unique filename
                    timestamp = now().strftime('%Y%m%d_%H%M%S')
                    new_filename = f"{user.employee_id}_{timestamp}.{file_ext}"
                    filepath = os.path.join(upload_folder, new_filename)
                    
                    # Save file
                    avatar_file.save(filepath)
                    
                    # Update user avatar_url
                    user.avatar_url = f"/static/uploads/avatars/{new_filename}"
                    print(f"[EDIT USER] Avatar uploaded: {new_filename}")
        
        print(f"[EDIT USER] Committing changes to database...")
        db.session.commit()
        print(f"[EDIT USER] ✓ Database commit successful")
        
        # Tự động đăng ký suất ăn CHỈ KHI user đang làm việc
        success_message = f'Đã cập nhật thông tin {user.name}'
        
        # Nếu nghỉ việc, thông báo đã xóa đăng ký
        if user.work_status == 'resigned':
            if deleted_count > 0:
                success_message += f' và xóa {deleted_count} đăng ký cơm trong tương lai'
        else:
            # Chỉ auto-register khi đang làm việc
            try:
                if user.work_status == 'working':
                    print(f"[EDIT USER] User is working, running auto-register...")
                    auto_result = auto_register_meals_for_user(user.id, days=30)
                    print(f"[EDIT USER] Auto-register result: {auto_result}")
                    if auto_result['success'] and auto_result['registered'] > 0:
                        success_message += f' và tự động đăng ký {auto_result["registered"]} ngày'
            except Exception as auto_error:
                print(f"[EDIT USER] Warning: Auto-register failed but user update succeeded: {auto_error}")
        
        print(f"[EDIT USER] ✓ Success: {success_message}\n")
        return {'success': True, 'message': success_message}
    
    except Exception as e:
        db.session.rollback()
        print(f"[EDIT USER] ❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'message': f'Lỗi: {str(e)}'}, 500


@app.route('/users/<int:user_id>/clear-avatar', methods=['POST'])
@login_required
def clear_avatar(user_id):
    """Clear user avatar"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    try:
        user = User.query.get_or_404(user_id)
        
        # Delete old avatar file if exists
        if user.avatar_url:
            import os
            old_file = user.avatar_url.lstrip('/')
            if os.path.exists(old_file):
                try:
                    os.remove(old_file)
                except Exception as e:
                    print(f"Error deleting file: {e}")
        
        # Clear avatar_url in database
        user.avatar_url = None
        db.session.commit()
        
        return {'success': True, 'message': 'Đã xóa ảnh đại diện'}
    
    except Exception as e:
        db.session.rollback()
        print(f"Error in clear_avatar: {e}")
        return {'success': False, 'message': f'Lỗi: {str(e)}'}, 500


@app.route('/admin/users/add', methods=['POST'])
@login_required
def add_user():
    """Add new user"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    from werkzeug.utils import secure_filename
    import os
    from datetime import datetime
    
    # Handle both JSON and FormData
    try:
        if request.is_json:
            data = request.json
        else:
            data = request.form.to_dict()
        
        # Validate required fields
        required_fields = ['employee_id', 'name', 'password', 'department']
        for field in required_fields:
            if field not in data or not data[field]:
                return {'success': False, 'message': f'Thiếu thông tin: {field}'}, 400
        
        # Check if employee_id already exists
        if User.query.filter_by(employee_id=data['employee_id']).first():
            return {'success': False, 'message': 'Mã nhân viên đã tồn tại'}, 400
        
        # Find department_id from department name
        from models import Department
        dept = Department.query.filter_by(name=data['department']).first()
        department_id = dept.id if dept else None
        
        # Create new user
        new_user = User(
            employee_id=data['employee_id'],
            name=data['name'],
            department_id=department_id,  # Department foreign key
            role=data.get('role', 'user'),
            work_status=data.get('work_status', 'working'),
            is_active=True,  # Mặc định luôn active khi tạo mới
            can_approve=data.get('can_approve') in ['true', 'True', True, 1, '1'] if 'can_approve' in data else False,
            can_register=data.get('can_register') in ['true', 'True', True, 1, '1'] if 'can_register' in data else True
        )
        new_user.set_password(data['password'])
        
        # Handle avatar upload
        if 'avatar' in request.files:
            avatar_file = request.files['avatar']
            if avatar_file and avatar_file.filename:
                # Validate file extension
                allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
                filename = secure_filename(avatar_file.filename)
                file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
                
                if file_ext in allowed_extensions:
                    # Create upload folder if not exists
                    upload_folder = os.path.join('static', 'uploads', 'avatars')
                    os.makedirs(upload_folder, exist_ok=True)
                    
                    # Generate unique filename
                    timestamp = now().strftime('%Y%m%d_%H%M%S')
                    new_filename = f"{new_user.employee_id}_{timestamp}.{file_ext}"
                    filepath = os.path.join(upload_folder, new_filename)
                    
                    # Save file
                    avatar_file.save(filepath)
                    
                    # Update user avatar_url
                    new_user.avatar_url = f"/static/uploads/avatars/{new_filename}"
        
        db.session.add(new_user)
        db.session.commit()
        
        # Tự động đăng ký suất ăn 30 ngày cho user mới
        auto_result = auto_register_meals_for_user(new_user.id, days=30)
        
        success_message = f'Đã thêm nhân viên {new_user.name}'
        if auto_result['success']:
            success_message += f' và tự động đăng ký {auto_result["registered"]} ngày'
        
        return {'success': True, 'message': success_message}
    
    except Exception as e:
        db.session.rollback()
        print(f"Error in add_user: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'message': f'Lỗi: {str(e)}'}, 500

@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@login_required
def reset_user_password(user_id):
    """Reset user password to default (123456)"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    user = User.query.get_or_404(user_id)
    
    # Reset password to default
    user.password = generate_password_hash('123456')
    db.session.commit()
    
    return {'success': True, 'message': f'Đã reset mật khẩu của {user.name} về 123456'}


@app.route('/admin/users/download-template')
@login_required
def download_users_template():
    """Download Excel template for importing users"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import io
    from flask import send_file
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users Template"
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Define headers
    headers = [
        "Mã nhân viên *",
        "Họ và tên *", 
        "Phòng ban *",
        "Mật khẩu",
        "Quyền",
        "Trạng thái",
        "Quyền phê duyệt",
        "Quyền đăng ký"
    ]
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add example data
    example_data = [
        ["10001", "Nguyễn Văn A", "Phòng IT", "123456", "user", "working", "false", "false"],
        ["10002", "Trần Thị B", "Phòng HR", "123456", "user", "working", "false", "false"],
        ["20001", "Lê Văn C", "Phòng Sản xuất", "123456", "user", "working", "false", "false"]
    ]
    
    for row_idx, row_data in enumerate(example_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add instructions sheet
    ws_instructions = wb.create_sheet("Hướng dẫn")
    instructions = [
        ["HƯỚNG DẪN IMPORT USERS"],
        [""],
        ["Các cột bắt buộc (có dấu *):"],
        ["- Mã nhân viên: Mã duy nhất của nhân viên (VD: 10001, 20001)"],
        ["- Họ và tên: Tên đầy đủ của nhân viên"],
        ["- Phòng ban: Tên phòng ban (nếu chưa có sẽ tự động tạo mới)"],
        [""],
        ["Các cột tùy chọn:"],
        ["- Mật khẩu: Mặc định là 123456 nếu để trống"],
        ["- Quyền: user hoặc admin (mặc định: user)"],
        ["- Trạng thái: working, business_trip, resigned (mặc định: working)"],
        ["- Quyền phê duyệt: true hoặc false (mặc định: false)"],
        ["- Quyền đăng ký: true hoặc false (mặc định: false)"],
        [""],
        ["Lưu ý:"],
        ["- Không xóa dòng tiêu đề"],
        ["- Mã nhân viên không được trùng"],
        ["- Phòng ban chưa tồn tại sẽ được tự động tạo mới"],
        ["- Xóa các dòng ví dụ trước khi import dữ liệu thật"]
    ]
    
    for row_idx, instruction in enumerate(instructions, start=1):
        ws_instructions.cell(row=row_idx, column=1, value=instruction[0])
        if row_idx == 1:
            ws_instructions.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    
    ws_instructions.column_dimensions['A'].width = 60
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'users_template_{today().strftime("%Y%m%d")}.xlsx'
    )




@app.route('/admin/users/check-exists', methods=['POST'])
@login_required
def check_users_exist():
    """Check if users exist by employee_id"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    try:
        data = request.get_json()
        employee_ids = data.get('employee_ids', [])
        
        if not employee_ids:
            return {'success': True, 'existing_ids': []}
        
        # Query existing users
        existing_users = User.query.filter(User.employee_id.in_(employee_ids)).all()
        existing_ids = [user.employee_id for user in existing_users]
        
        return {
            'success': True,
            'existing_ids': existing_ids
        }
    except Exception as e:
        print(f"Error checking users: {str(e)}")
        return {'success': False, 'message': str(e)}, 500


@app.route('/admin/users/import', methods=['POST', 'OPTIONS'])
@login_required
def import_users_excel():
    """Import users from Excel file"""
    # Handle OPTIONS request for CORS
    if request.method == 'OPTIONS':
        return '', 204
    
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    # Debug logging
    print(f"\n=== IMPORT REQUEST DEBUG ===")
    print(f"Request method: {request.method}")
    print(f"Request content type: {request.content_type}")
    print(f"Request files keys: {list(request.files.keys())}")
    print(f"Request form keys: {list(request.form.keys())}")
    print(f"Request files: {request.files}")
    print(f"Request data: {request.data[:100] if request.data else 'No data'}")
    
    if 'file' not in request.files:
        print("ERROR: 'file' not in request.files")
        print(f"Available keys: {list(request.files.keys())}")
        return {'success': False, 'message': 'Không có file được tải lên'}, 400
    
    file = request.files['file']
    print(f"File received: {file.filename}")
    
    if file.filename == '':
        print("ERROR: file.filename is empty")
        return {'success': False, 'message': 'Không có file được chọn'}, 400
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        print(f"ERROR: Invalid file extension: {file.filename}")
        return {'success': False, 'message': 'File phải là định dạng Excel (.xlsx hoặc .xls)'}, 400
    
    print(f"File validation passed, starting import...")
    print(f"=== END DEBUG ===\n")
    
    import openpyxl
    from models import Department
    
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        # Skip header row
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        if not rows:
            return {'success': False, 'message': 'File Excel không có dữ liệu'}, 400
        
        print(f"[IMPORT] Total rows to process: {len(rows)}")
        
        success_count = 0
        error_count = 0
        errors = []
        
        for row_idx, row in enumerate(rows, start=2):
            try:
                # Skip empty rows
                if not any(row):
                    print(f"[IMPORT] Row {row_idx}: Empty row, skipping")
                    continue
                
                print(f"[IMPORT] Processing row {row_idx}: {row[:3]}")  # Print first 3 columns
                
                employee_id = str(row[0]).strip() if row[0] else None
                name = str(row[1]).strip() if row[1] else None
                department_name = str(row[2]).strip() if row[2] else None
                password = str(row[3]).strip() if row[3] else '123456'
                role = str(row[4]).strip().lower() if row[4] else 'user'
                work_status = str(row[5]).strip().lower() if row[5] else 'working'
                can_approve = str(row[6]).strip().lower() == 'true' if row[6] else False
                can_register = str(row[7]).strip().lower() == 'true' if row[7] else False
                
                print(f"[IMPORT] Row {row_idx}: employee_id={employee_id}, name={name}, dept={department_name}")
                
                # Validate required fields
                if not employee_id or not name or not department_name:
                    error_msg = f"Dòng {row_idx}: Thiếu thông tin bắt buộc (Mã NV, Tên, Phòng ban)"
                    print(f"[IMPORT] {error_msg}")
                    errors.append(error_msg)
                    error_count += 1
                    continue
                
                # Find or create department
                department = Department.query.filter_by(name=department_name).first()
                if not department:
                    # Auto-create department if not exists
                    department = Department(
                        code=department_name.upper().replace(' ', '_')[:50],  # Generate code from name
                        name=department_name,
                        description=f'Tự động tạo từ import'
                    )
                    db.session.add(department)
                    db.session.flush()  # Get department.id without committing
                
                # Validate role
                if role not in ['user', 'admin']:
                    role = 'user'
                
                # Validate work_status
                if work_status not in ['working', 'business_trip', 'resigned']:
                    work_status = 'working'
                
                # Check if user already exists - UPDATE instead of skip
                existing_user = User.query.filter_by(employee_id=employee_id).first()
                if existing_user:
                    # UPDATE existing user
                    existing_user.name = name
                    existing_user.department_id = department.id
                    existing_user.role = role
                    existing_user.work_status = work_status
                    existing_user.can_approve = can_approve
                    existing_user.can_register = can_register
                    # Only update password if provided and not default
                    if password and password != '123456':
                        existing_user.set_password(password)
                    success_count += 1
                    print(f"[IMPORT] Updated user: {employee_id}")
                else:
                    # CREATE new user
                    new_user = User(
                        employee_id=employee_id,
                        name=name,
                        department_id=department.id,
                        role=role,
                        work_status=work_status,
                        is_active=True,
                        can_approve=can_approve,
                        can_register=can_register
                    )
                    new_user.set_password(password)
                    db.session.add(new_user)
                    success_count += 1
                    print(f"[IMPORT] Created user: {employee_id}")
                
            except Exception as e:
                error_msg = f"Dòng {row_idx}: Lỗi - {str(e)}"
                print(f"[IMPORT] {error_msg}")
                import traceback
                print(traceback.format_exc())
                errors.append(error_msg)
                error_count += 1
                continue
        
        print(f"[IMPORT] Finished processing. Success: {success_count}, Errors: {error_count}")
        
        # Commit all changes
        if success_count > 0:
            db.session.commit()
            print(f"[IMPORT] Committed {success_count} changes to database")
        else:
            print(f"[IMPORT] No changes to commit")
        
        # Prepare response message
        message = f"Import thành công {success_count} user"
        if error_count > 0:
            message += f", {error_count} lỗi"
        
        print(f"[IMPORT] Response: {message}")
        
        return {
            'success': True,
            'message': message,
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors[:10]  # Limit to first 10 errors
        }, 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"Import error: {str(e)}")
        print(traceback.format_exc())
        return {'success': False, 'message': f'Lỗi khi đọc file: {str(e)}'}, 500




@app.route('/admin/departments')
@login_required
def admin_departments():
    """Admin departments management page"""
    if current_user.role != 'admin':
        flash('Không có quyền truy cập', 'error')
        return redirect(url_for('dashboard'))
    
    from models import Department
    
    # Get all departments
    departments = Department.query.order_by(Department.name.asc()).all()
    
    # Create list with empty manager info for template compatibility
    departments_with_manager = [(dept, None) for dept in departments]
    
    return render_template('admin_departments.html', departments=departments_with_manager)

@app.route('/admin/departments/add', methods=['POST'])
@login_required
def add_department():
    """Add new department"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    from models import Department
    
    try:
        if request.is_json:
            data = request.json
        else:
            data = request.form.to_dict()
        
        # Validate required fields
        required_fields = ['code', 'name']
        for field in required_fields:
            if field not in data or not data[field]:
                return {'success': False, 'message': f'Thiếu thông tin: {field}'}, 400
        
        # Check if code already exists
        if Department.query.filter_by(code=data['code']).first():
            return {'success': False, 'message': 'Mã phòng ban đã tồn tại'}, 400
        
        # Create new department
        new_department = Department(
            code=data['code'].upper(),
            name=data['name'],
            description=data.get('description', '')
        )
        
        db.session.add(new_department)
        db.session.commit()
        
        return {'success': True, 'message': f'Đã thêm phòng ban {data["name"]}'}
        
    except Exception as e:
        db.session.rollback()
        return {'success': False, 'message': f'Lỗi: {str(e)}'}, 500

@app.route('/admin/departments/<int:department_id>/edit', methods=['POST'])
@login_required
def edit_department(department_id):
    """Edit department"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    from models import Department
    
    department = Department.query.get_or_404(department_id)
    
    try:
        if request.is_json:
            data = request.json
        else:
            data = request.form.to_dict()
        
        if 'code' in data:
            # Check if code already exists (excluding current department)
            existing = Department.query.filter(Department.code == data['code'], Department.id != department_id).first()
            if existing:
                return {'success': False, 'message': 'Mã phòng ban đã tồn tại'}, 400
            department.code = data['code'].upper()
        
        if 'name' in data:
            department.name = data['name']
        if 'description' in data:
            department.description = data['description']
        
        department.updated_at = datetime.utcnow()
        db.session.commit()
        
        return {'success': True, 'message': f'Đã cập nhật phòng ban {department.name}'}
        
    except Exception as e:
        db.session.rollback()
        return {'success': False, 'message': f'Lỗi: {str(e)}'}, 500

@app.route('/admin/departments/<int:department_id>/delete', methods=['POST'])
@login_required
def delete_department(department_id):
    """Delete department"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    from models import Department
    
    department = Department.query.get_or_404(department_id)
    department_name = department.name
    
    try:
        # Check if department has users
        users_count = User.query.filter_by(department_id=department_id).count()
        if users_count > 0:
            return {'success': False, 'message': f'Không thể xóa phòng ban có {users_count} nhân viên'}, 400
        
        db.session.delete(department)
        db.session.commit()
        
        return {'success': True, 'message': f'Đã xóa phòng ban {department_name}'}
        
    except Exception as e:
        db.session.rollback()
        return {'success': False, 'message': f'Lỗi: {str(e)}'}, 500

# ============= END DEPARTMENT MANAGEMENT ROUTES =============

# ============= DEPARTMENT MANAGER ROUTES - REMOVED =============
# Old department_managers system has been replaced with level-based hierarchy
# No longer need to manually manage department managers
# Users with lower level automatically become approvers for their subordinates
# ============= END DEPARTMENT MANAGER ROUTES =============

# ============= OVERTIME APPROVAL ROUTES =============
@app.route('/admin/overtime-approvals')
@login_required
def admin_overtime_approvals():
    """Admin overtime approvals page"""
    if current_user.role != 'admin':
        flash('Bạn không có quyền truy cập trang này', 'error')
        return redirect(url_for('dashboard'))
    
    return render_template('admin_overtime_approvals.html')

@app.route('/admin/overtime-requests/list')
@login_required
def list_overtime_requests():
    """Get all overtime requests for admin"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Không có quyền'}), 403
    
    try:
        status_filter = request.args.get('status', 'all')
        
        query = OvertimeRequest.query
        if status_filter != 'all':
            query = query.filter_by(status=status_filter)
        
        requests = query.order_by(OvertimeRequest.created_at.desc()).all()
        
        result_list = []
        for r in requests:
            try:
                # Get requester info
                requester = None
                requester_name = ''
                requester_employee_id = ''
                if r.user_id:
                    try:
                        requester = User.query.get(r.user_id)
                        if requester:
                            requester_name = requester.name or ''
                            requester_employee_id = requester.employee_id or ''
                    except Exception as e:
                        print(f"Error getting requester {r.user_id}: {e}")
                
                # Get approver info
                approver = None
                approver_name = ''
                approver_employee_id = ''
                if r.manager_id:
                    try:
                        approver = User.query.get(r.manager_id)
                        if approver:
                            approver_name = approver.name or ''
                            approver_employee_id = approver.employee_id or ''
                    except Exception as e:
                        print(f"Error getting approver {r.manager_id}: {e}")
                
                result_list.append({
                    'id': r.id,
                    'employee_name': r.employee_name or '',
                    'employee_id': r.employee_id or '',
                    'department': r.department or '',
                    'overtime_date': r.overtime_date.strftime('%Y-%m-%d') if r.overtime_date else '',
                    'start_time': r.start_time.strftime('%H:%M') if r.start_time else '',
                    'end_time': r.end_time.strftime('%H:%M') if r.end_time else '',
                    'total_hours': float(r.total_hours) if r.total_hours else 0,
                    'reason': r.reason or '',
                    'status': r.status or 'pending',
                    'manager_comment': r.manager_comment or '',
                    'created_at': r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
                    # Add requester info - use employee_id as username
                    'requester_name': requester_name,
                    'requester_username': requester_employee_id,
                    # Add approver info - use employee_id as username
                    'approver_name': approver_name,
                    'approver_username': approver_employee_id
                })
            except Exception as e:
                print(f"Error processing overtime request {r.id}: {e}")
                import traceback
                traceback.print_exc()
                continue
        
        return jsonify({
            'success': True,
            'requests': result_list
        })
    except Exception as e:
        print(f"Error in list_overtime_requests: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

# ============= END OVERTIME APPROVAL ROUTES =============

@app.route('/admin/overtime-requests/<int:request_id>/delete', methods=['POST'])
@login_required
def delete_overtime_request(request_id):
    """Delete an overtime request (admin only)"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Bạn không có quyền xóa yêu cầu này'}), 403
    
    try:
        overtime_request = OvertimeRequest.query.get(request_id)
        if not overtime_request:
            return jsonify({'success': False, 'message': 'Không tìm thấy yêu cầu'}), 404
        
        db.session.delete(overtime_request)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Đã xóa yêu cầu thành công'})
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting overtime request: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'Có lỗi xảy ra khi xóa yêu cầu'}), 500

# ============= MANAGER OVERTIME APPROVAL ROUTES =============
@app.route('/manager/overtime-approvals')
@login_required
def manager_overtime_approvals():
    """Manager overtime approvals page"""
    # Admin always has access
    if current_user.role == 'admin':
        return render_template('manager_overtime_approvals.html', is_manager=True)
    
    # Check if user has approval permission
    if current_user.can_approve:
        return render_template('manager_overtime_approvals.html', is_manager=True)
    
    # Or if user has subordinates
    if current_user.is_manager():
        return render_template('manager_overtime_approvals.html', is_manager=True)
    
    flash('Bạn không có quyền truy cập trang này', 'error')
    return redirect(url_for('dashboard'))


@app.route('/manager/overtime-requests')
@login_required
def get_manager_overtime_requests():
    """Get overtime requests that this user can approve (based on level hierarchy)"""
    try:
        # Admin always has access
        if current_user.role == 'admin':
            # Admin can see all requests
            requests = OvertimeRequest.query.order_by(
                OvertimeRequest.created_at.desc()
            ).all()
        else:
            # Check if user has approval permission
            can_approve = current_user.can_approve
            
            if not can_approve:
                return jsonify({'success': False, 'message': 'Không có quyền'}), 403
            
            # Get subordinate user IDs
            subordinates = current_user.get_subordinates()
            subordinate_ids = [sub.id for sub in subordinates]
            
            # If has approval permission, include own requests
            if current_user.can_approve:
                subordinate_ids.append(current_user.id)
            
            # Get requests from subordinates (and self if Trưởng Phòng)
            requests = OvertimeRequest.query.filter(
                OvertimeRequest.user_id.in_(subordinate_ids)
            ).order_by(OvertimeRequest.created_at.desc()).all() if subordinate_ids else []
        
        result_list = []
        for r in requests:
            # Get requester info
            requester = None
            requester_name = ''
            requester_employee_id = ''
            if r.user_id:
                try:
                    requester = User.query.get(r.user_id)
                    if requester:
                        requester_name = requester.name or ''
                        requester_employee_id = requester.employee_id or ''
                except Exception as e:
                    print(f"Error getting requester {r.user_id}: {e}")
            
            # Get approver info
            approver = None
            approver_name = ''
            approver_employee_id = ''
            if r.manager_id:
                try:
                    approver = User.query.get(r.manager_id)
                    if approver:
                        approver_name = approver.name or ''
                        approver_employee_id = approver.employee_id or ''
                except Exception as e:
                    print(f"Error getting approver {r.manager_id}: {e}")
            
            result_list.append({
                'id': r.id,
                'employee_name': r.employee_name,
                'employee_id': r.employee_id,
                'department': r.department,
                'overtime_date': r.overtime_date.strftime('%Y-%m-%d'),
                'start_time': r.start_time.strftime('%H:%M') if r.start_time else '00:00',
                'end_time': r.end_time.strftime('%H:%M') if r.end_time else '00:00',
                'total_hours': float(r.total_hours) if r.total_hours else 0,
                'reason': r.reason or '',
                'status': r.status,
                'manager_comment': r.manager_comment or '',
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
                # Add requester info
                'requester_name': requester_name,
                'requester_username': requester_employee_id,
                # Add approver info
                'approver_name': approver_name,
                'approver_username': approver_employee_id
            })
        
        return jsonify({
            'success': True,
            'requests': result_list
        })
    except Exception as e:
        print(f"Error in get_manager_overtime_requests: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/manager/overtime-requests/<int:request_id>/approve', methods=['POST'])
@login_required
def approve_overtime_request(request_id):
    """Approve overtime request (level-based authorization)"""
    overtime_request = OvertimeRequest.query.get_or_404(request_id)
    requester = User.query.get(overtime_request.user_id)
    
    # Check authorization
    can_approve = False
    
    # Admin can approve all
    if current_user.role == 'admin':
        can_approve = True
    # Can approve subordinates
    elif current_user.can_approve_for(requester):
        can_approve = True
    # Can approve own request if has approval permission
    elif requester.id == current_user.id and current_user.can_approve:
        can_approve = True
    
    if not can_approve:
        return {'success': False, 'message': 'Bạn không có quyền duyệt yêu cầu này'}, 403
    
    # Check if already processed
    if overtime_request.status != 'pending':
        return {'success': False, 'message': 'Yêu cầu đã được xử lý'}, 400
    
    data = request.json
    comment = data.get('comment', '').strip()
    
    overtime_request.status = 'approved'
    overtime_request.manager_id = current_user.id
    overtime_request.manager_approved_at = datetime.utcnow()
    if requester.id == current_user.id:
        overtime_request.manager_comment = 'Tự duyệt (cấp cao nhất trong phòng)'
    else:
        overtime_request.manager_comment = comment if comment else None
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'Đã duyệt yêu cầu tăng ca của {overtime_request.employee_name}'})

@app.route('/manager/overtime-requests/<int:request_id>/reject', methods=['POST'])
@login_required
def reject_overtime_request(request_id):
    """Reject overtime request (level-based authorization)"""
    overtime_request = OvertimeRequest.query.get_or_404(request_id)
    requester = User.query.get(overtime_request.user_id)
    
    # Check authorization
    can_reject = False
    
    # Admin can reject all
    if current_user.role == 'admin':
        can_reject = True
    # Can reject subordinates
    elif current_user.can_approve_for(requester):
        can_reject = True
    # Can reject own request if has approval permission
    elif requester.id == current_user.id and current_user.can_approve:
        can_reject = True
    
    if not can_reject:
        return {'success': False, 'message': 'Bạn không có quyền xử lý yêu cầu này'}, 403
    
    # Check if already processed
    if overtime_request.status != 'pending':
        return {'success': False, 'message': 'Yêu cầu đã được xử lý'}, 400
    
    data = request.json
    comment = data.get('comment', '').strip()
    
    if not comment:
        return {'success': False, 'message': 'Vui lòng nhập lý do từ chối'}, 400
    
    overtime_request.status = 'rejected'
    overtime_request.manager_id = current_user.id
    overtime_request.manager_approved_at = datetime.utcnow()
    overtime_request.manager_comment = comment
    
    db.session.commit()
    
    return jsonify({'success': True, 'message': f'Đã từ chối yêu cầu tăng ca của {overtime_request.employee_name}'})

@app.route('/manager/overtime-requests/<int:request_id>/update-count', methods=['POST'])
@login_required
def update_overtime_count(request_id):
    """Update number of people in overtime request"""
    overtime_request = OvertimeRequest.query.get_or_404(request_id)
    requester = User.query.get(overtime_request.user_id)
    
    # Check authorization - only managers can update
    can_update = False
    
    if current_user.role == 'admin':
        can_update = True
    elif current_user.can_approve_for(requester):
        can_update = True
    elif requester.id == current_user.id and current_user.can_approve:
        can_update = True
    
    if not can_update:
        return {'success': False, 'message': 'Bạn không có quyền cập nhật'}, 403
    
    # Only allow updating pending requests
    if overtime_request.status != 'pending':
        return {'success': False, 'message': 'Chỉ có thể cập nhật yêu cầu đang chờ duyệt'}, 400
    
    data = request.json
    new_count = data.get('total_hours')
    
    if not new_count or int(new_count) < 1:
        return {'success': False, 'message': 'Số người phải lớn hơn 0'}, 400
    
    overtime_request.total_hours = int(new_count)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Đã cập nhật số người'})

# ============= END MANAGER OVERTIME APPROVAL ROUTES =============

# ============= OVERTIME REPORTS ROUTES =============
@app.route('/overtime/reports')
@login_required
def overtime_reports():
    """Overtime reports page - admin only"""
    if current_user.role != 'admin':
        flash('Bạn không có quyền truy cập trang này', 'error')
        return redirect(url_for('dashboard'))
    
    return render_template('overtime_reports.html')

@app.route('/overtime/reports/data')
@login_required
def get_overtime_reports_data():
    """Get overtime reports data"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    from sqlalchemy import func, extract
    from datetime import datetime, timedelta
    
    period_type = request.args.get('period_type', 'month')
    period_value = int(request.args.get('period_value', now().month))
    year = int(request.args.get('year', now().year))
    
    # Calculate date range
    if period_type == 'month':
        start_date = datetime(year, period_value, 1).date()
        if period_value == 12:
            end_date = datetime(year, 12, 31).date()
        else:
            end_date = (datetime(year, period_value + 1, 1) - timedelta(days=1)).date()
    elif period_type == 'quarter':
        start_month = (period_value - 1) * 3 + 1
        start_date = datetime(year, start_month, 1).date()
        end_month = start_month + 2
        if end_month == 12:
            end_date = datetime(year, 12, 31).date()
        else:
            end_date = (datetime(year, end_month + 1, 1) - timedelta(days=1)).date()
    else:  # year
        start_date = datetime(year, 1, 1).date()
        end_date = datetime(year, 12, 31).date()
    
    # Get approved requests in period
    requests = OvertimeRequest.query.filter(
        OvertimeRequest.overtime_date >= start_date,
        OvertimeRequest.overtime_date <= end_date,
        OvertimeRequest.status == 'approved'
    ).all()
    
    # Calculate summary
    total_hours = sum(float(r.total_hours) for r in requests)
    unique_employees = len(set(r.employee_id for r in requests))
    total_requests = len(requests)
    avg_hours = total_hours / unique_employees if unique_employees > 0 else 0
    
    # Group by department
    dept_data = {}
    for r in requests:
        if r.department not in dept_data:
            dept_data[r.department] = 0
        dept_data[r.department] += float(r.total_hours)
    
    by_department = [
        {'department': dept, 'total_hours': hours}
        for dept, hours in sorted(dept_data.items(), key=lambda x: x[1], reverse=True)
    ]
    
    # Trend data
    trend_data = []
    if period_type == 'month':
        # Daily trend
        current_date = start_date
        while current_date <= end_date:
            day_requests = [r for r in requests if r.overtime_date == current_date]
            day_hours = sum(float(r.total_hours) for r in day_requests)
            trend_data.append({
                'label': f'{current_date.day}/{current_date.month}',
                'total_hours': day_hours
            })
            current_date += timedelta(days=1)
    elif period_type == 'quarter':
        # Monthly trend
        for month in range(start_month, end_month + 1):
            month_requests = [r for r in requests if r.overtime_date.month == month]
            month_hours = sum(float(r.total_hours) for r in month_requests)
            trend_data.append({
                'label': f'T{month}',
                'total_hours': month_hours
            })
    else:
        # Quarterly trend
        for q in range(1, 5):
            q_start_month = (q - 1) * 3 + 1
            q_end_month = q_start_month + 2
            q_requests = [r for r in requests if q_start_month <= r.overtime_date.month <= q_end_month]
            q_hours = sum(float(r.total_hours) for r in q_requests)
            trend_data.append({
                'label': f'Q{q}',
                'total_hours': q_hours
            })
    
    # Top employees
    emp_data = {}
    for r in requests:
        key = r.employee_id
        if key not in emp_data:
            emp_data[key] = {
                'employee_id': r.employee_id,
                'employee_name': r.employee_name,
                'department': r.department,
                'total_hours': 0,
                'request_count': 0
            }
        emp_data[key]['total_hours'] += float(r.total_hours)
        emp_data[key]['request_count'] += 1
    
    top_employees = sorted(emp_data.values(), key=lambda x: x['total_hours'], reverse=True)[:10]
    
    return {
        'success': True,
        'summary': {
            'total_hours': total_hours,
            'total_employees': unique_employees,
            'total_requests': total_requests,
            'avg_hours': avg_hours
        },
        'by_department': by_department,
        'trend': trend_data,
        'top_employees': top_employees
    }

@app.route('/overtime/reports/export')
@login_required
def export_overtime_reports():
    """Export overtime reports to Excel"""
    if current_user.role != 'admin':
        flash('Bạn không có quyền', 'error')
        return redirect(url_for('dashboard'))
    
    from datetime import datetime, timedelta
    import io
    from flask import send_file
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        flash('Vui lòng cài đặt openpyxl: pip install openpyxl', 'error')
        return redirect(url_for('overtime_reports'))
    
    period_type = request.args.get('period_type', 'month')
    period_value = int(request.args.get('period_value', now().month))
    year = int(request.args.get('year', now().year))
    
    # Calculate date range
    if period_type == 'month':
        start_date = datetime(year, period_value, 1).date()
        if period_value == 12:
            end_date = datetime(year, 12, 31).date()
        else:
            end_date = (datetime(year, period_value + 1, 1) - timedelta(days=1)).date()
        period_name = f"Tháng {period_value}/{year}"
    elif period_type == 'quarter':
        start_month = (period_value - 1) * 3 + 1
        start_date = datetime(year, start_month, 1).date()
        end_month = start_month + 2
        if end_month == 12:
            end_date = datetime(year, 12, 31).date()
        else:
            end_date = (datetime(year, end_month + 1, 1) - timedelta(days=1)).date()
        period_name = f"Quý {period_value}/{year}"
    else:
        start_date = datetime(year, 1, 1).date()
        end_date = datetime(year, 12, 31).date()
        period_name = f"Năm {year}"
    
    # Get data
    requests = OvertimeRequest.query.filter(
        OvertimeRequest.overtime_date >= start_date,
        OvertimeRequest.overtime_date <= end_date,
        OvertimeRequest.status == 'approved'
    ).order_by(OvertimeRequest.overtime_date.desc()).all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo cáo tăng ca"
    
    # Header
    ws['A1'] = f"BÁO CÁO TĂNG CA - {period_name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Column headers
    headers = ['STT', 'Mã NV', 'Họ tên', 'Phòng ban', 'Ngày', 'Giờ bắt đầu', 'Giờ kết thúc', 'Tổng giờ']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for idx, req in enumerate(requests, 1):
        ws.cell(row=idx+3, column=1, value=idx)
        ws.cell(row=idx+3, column=2, value=req.employee_id)
        ws.cell(row=idx+3, column=3, value=req.employee_name)
        ws.cell(row=idx+3, column=4, value=req.department)
        ws.cell(row=idx+3, column=5, value=req.overtime_date.strftime('%d/%m/%Y'))
        ws.cell(row=idx+3, column=6, value=req.start_time.strftime('%H:%M'))
        ws.cell(row=idx+3, column=7, value=req.end_time.strftime('%H:%M'))
        ws.cell(row=idx+3, column=8, value=float(req.total_hours))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"BaoCaoTangCa_{period_name.replace('/', '_').replace(' ', '_')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# ============= END OVERTIME REPORTS ROUTES =============

# ============= MEAL MANAGEMENT ROUTES =============
@app.route('/admin/meals')
@login_required
def admin_meals():
    """Admin meal management page"""
    if current_user.role != 'admin':
        flash('Không có quyền truy cập', 'error')
        return redirect(url_for('dashboard'))
    
    from datetime import datetime, timedelta
    
    # Get current week (Monday to Sunday)
    today_date = today()
    start_of_week = today_date - timedelta(days=today_date.weekday())  # Monday
    
    days_of_week = []
    day_names = ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN']
    
    for i in range(7):
        day = start_of_week + timedelta(days=i)
        days_of_week.append({
            'date': day.strftime('%Y-%m-%d'),
            'day_name': day_names[i],
            'day_num': day.day
        })
    
    return render_template('admin_meals.html', 
                         days_of_week=days_of_week,
                         current_lang=session.get('lang', 'vi'),
                         t=get_translation())

@app.route('/admin/stats')
@login_required
def admin_stats():
    """Admin statistics and reports page"""
    if current_user.role != 'admin':
        flash('Không có quyền truy cập', 'error')
        return redirect(url_for('dashboard'))
    
    # TẮT auto-register khi vào trang thống kê để tránh chạy quá nhiều lần
    # Admin có thể dùng nút "Chạy Auto 30 ngày" nếu cần
    # try:
    #     print("\n[ADMIN STATS] Tự động chạy auto-register 30 ngày...")
    #     auto_register_meals_for_30_days()
    # except Exception as e:
    #     print(f"[ADMIN STATS] Lỗi khi chạy auto-register: {str(e)}")
    
    return render_template('admin_stats.html')


@app.route('/admin/auto-register/run', methods=['POST'])
@login_required
def run_auto_register():
    """Manually trigger auto-register for 30 days"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Không có quyền'}), 403
    
    try:
        print("\n[MANUAL TRIGGER] Admin chạy thủ công auto-register 30 ngày...")
        auto_register_meals_for_30_days()
        return jsonify({'success': True, 'message': 'Đã chạy tự động đăng ký 30 ngày thành công! Xem console để biết chi tiết.'})
    except Exception as e:
        print(f"[MANUAL TRIGGER] Lỗi: {str(e)}")
        return jsonify({'success': False, 'message': f'Lỗi: {str(e)}'}), 500


@app.route('/admin/stats/data')
@login_required
def get_stats_data():
    """Get meal statistics data"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    from datetime import datetime, timedelta
    from sqlalchemy import func, or_
    from models import Menu
    
    # Lấy date range từ request parameters (từ bộ lọc người dùng chọn)
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    today_date = today()
    
    # Nếu có date_from và date_to từ bộ lọc, dùng chúng
    if date_from and date_to:
        try:
            start_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            end_date = datetime.strptime(date_to, '%Y-%m-%d').date()
        except:
            # Nếu parse lỗi, dùng tuần hiện tại
            start_date = today_date - timedelta(days=today_date.weekday())
            end_date = start_date + timedelta(days=6)
    else:
        # Nếu không có bộ lọc, mặc định lấy tuần hiện tại
        start_date = today_date - timedelta(days=today_date.weekday())
        end_date = start_date + timedelta(days=6)
    
    period = request.args.get('period', 'custom')  # Để biết loại period cho chart
    
    # AUTO-REGISTER đã được tắt - sử dụng nút "Chạy Auto 30 ngày" hoặc đợi 16:00
    # Hoặc auto-register sẽ chạy khi tạo/edit user
    
    # Get registrations for the period with proper joins - only working users with employee_id starting with 10 or 20
    registrations = MealRegistration.query.join(
        User,
        MealRegistration.user_id == User.id
    ).join(
        Menu,
        MealRegistration.meal_id == Menu.id
    ).filter(
        MealRegistration.date >= start_date,
        MealRegistration.date <= end_date,
        MealRegistration.has_meal == True,
        User.work_status == 'working',  # CHỈ lấy user đang làm việc
        or_(
            User.employee_id.like('10%'),
            User.employee_id.like('20%')
        )
    ).all()
    
    # Debug: Found registrations for period
    
    # Calculate summary
    total_meals = len(registrations)
    normal_meals = sum(1 for r in registrations if r.menu and not r.menu.is_special)
    special_meals = sum(1 for r in registrations if r.menu and r.menu.is_special)
    
    # Get daily breakdown for chart (for week period)
    daily_data = []
    if period == 'week':
        for i in range(7):
            day = start_date + timedelta(days=i)
            day_regs = [r for r in registrations if r.date == day]
            normal = sum(1 for r in day_regs if r.menu and not r.menu.is_special)
            special = sum(1 for r in day_regs if r.menu and r.menu.is_special)
            
            day_names = ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN']
            daily_data.append({
                'day': day_names[i],
                'date': day.strftime('%d/%m'),
                'normal': normal,
                'special': special
            })
    
    # Get detailed registration list with user info
    registration_list = []
    for reg in registrations:
        try:
            registration_list.append({
                'id': reg.id,
                'user_name': reg.user.name,
                'employee_id': reg.user.employee_id,
                'department': reg.user.dept.name if reg.user.dept else 'N/A',  # Use dept relationship
                'date': reg.date.strftime('%d/%m/%Y'),
                'meal_name': reg.menu.dish_name if reg.menu else 'N/A',
                'meal_type': 'Cải Thiện' if (reg.menu and reg.menu.is_special) else 'Bình Thường',
                'is_special': reg.menu.is_special if reg.menu else False,
                'is_confirmed': reg.is_confirmed or False,
                'notes': reg.notes or ''
            })
        except Exception as e:
            # Error processing registration
            continue
    
    return jsonify({
        'success': True,
        'period': period,
        'start_date': start_date.strftime('%d/%m/%Y'),
        'end_date': end_date.strftime('%d/%m/%Y'),
        'summary': {
            'total': total_meals,
            'normal': normal_meals,
            'special': special_meals
        },
        'daily_data': daily_data,
        'registrations': registration_list
    })

@app.route('/admin/meal-registrations/list')
def list_meal_registrations():
    """Get meal registrations list with confirmation status"""
    # Skip authentication for API testing
    # if current_user.role != 'admin':
    #     return jsonify({'success': False, 'message': 'Không có quyền'}), 403
    
    try:
        from datetime import datetime, timedelta
        from sqlalchemy import or_
        from models import Menu  # Import Menu model
        
        # Get filter parameters
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        meal_type_filter = request.args.get('meal_type', 'all')  # all, normal, special
        status_filter = request.args.get('status', 'all')  # all, confirmed, unconfirmed
        
        # Build query - specify join condition explicitly
        query = MealRegistration.query.join(
            User, 
            MealRegistration.user_id == User.id
        ).join(
            Menu,
            MealRegistration.meal_id == Menu.id
        ).filter(
            MealRegistration.has_meal == True,
            or_(
                User.employee_id.like('10%'),
                User.employee_id.like('20%')
            )
        )
        
        # Apply date filters
        if date_from:
            query = query.filter(MealRegistration.date >= datetime.strptime(date_from, '%Y-%m-%d').date())
        if date_to:
            query = query.filter(MealRegistration.date <= datetime.strptime(date_to, '%Y-%m-%d').date())
        
        # Apply meal type filter
        if meal_type_filter == 'special':
            query = query.filter(Menu.is_special == True)
        elif meal_type_filter == 'normal':
            query = query.filter(Menu.is_special == False)
        
        # Apply status filter (only for special meals)
        if status_filter == 'confirmed':
            query = query.filter(MealRegistration.is_confirmed == True)
        elif status_filter == 'unconfirmed':
            query = query.filter(MealRegistration.is_confirmed == False)
        
        registrations = query.order_by(MealRegistration.date.desc()).all()
        
        # Build response
        result_list = []
        for reg in registrations:
            try:
                # Get confirmer info if confirmed
                confirmer_name = ''
                confirmer_employee_id = ''
                if reg.confirmed_by:
                    confirmer = User.query.get(reg.confirmed_by)
                    if confirmer:
                        confirmer_name = confirmer.name or ''
                        confirmer_employee_id = confirmer.employee_id or ''
                
                result_list.append({
                    'id': reg.id,
                    'user_name': reg.user.name,
                    'employee_id': reg.user.employee_id,
                    'department': reg.user.dept.name if reg.user.dept else 'N/A',
                    'date': reg.date.strftime('%Y-%m-%d'),
                    'meal_name': reg.menu.dish_name if reg.menu else 'N/A',
                    'meal_type': 'Cải Thiện' if (reg.menu and reg.menu.is_special) else 'Bình Thường',
                    'is_special': reg.menu.is_special if reg.menu else False,
                    'is_confirmed': reg.is_confirmed or False,
                    'confirmed_at': reg.confirmed_at.strftime('%Y-%m-%d %H:%M') if reg.confirmed_at else '',
                    'confirmer_name': confirmer_name,
                    'confirmer_username': confirmer_employee_id,
                    'notes': reg.notes or ''
                })
            except Exception as e:
                print(f"Error processing registration {reg.id}: {e}")
                continue
        
        return jsonify({
            'success': True,
            'registrations': result_list
        })
    except Exception as e:
        print(f"Error in list_meal_registrations: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/meal-registrations/<int:registration_id>/confirm', methods=['POST'])
def confirm_meal_registration(registration_id):
    """Confirm meal registration (for special meals)"""
    # Skip authentication for API testing
    # if current_user.role != 'admin':
    #     return {'success': False, 'message': 'Không có quyền'}, 403
    
    try:
        registration = MealRegistration.query.get_or_404(registration_id)
        
        # Check if it's a special meal
        if not registration.menu or not registration.menu.is_special:
            return jsonify({'success': False, 'message': 'Chỉ có thể xác nhận suất cơm cải thiện'}), 400
        
        # Update confirmation status
        registration.is_confirmed = True
        registration.confirmed_at = datetime.utcnow()
        # registration.confirmed_by = current_user.id  # Skip for testing
        registration.confirmed_by = 1  # Use admin ID for testing
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Đã xác nhận hoàn thành suất cơm cải thiện cho {registration.user.name}'
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error in confirm_meal_registration: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/meal-registrations/<int:registration_id>/unconfirm', methods=['POST'])
def unconfirm_meal_registration(registration_id):
    """Unconfirm meal registration"""
    # Skip authentication for API testing
    # if current_user.role != 'admin':
    #     return {'success': False, 'message': 'Không có quyền'}, 403
    
    try:
        registration = MealRegistration.query.get_or_404(registration_id)
        
        # Update confirmation status
        registration.is_confirmed = False
        registration.confirmed_at = None
        registration.confirmed_by = None
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Đã hủy xác nhận cho {registration.user.name}'
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error in unconfirm_meal_registration: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/meal-registrations/bulk-confirm', methods=['POST'])
def bulk_confirm_meal_registrations():
    """Bulk confirm multiple meal registrations"""
    # Skip authentication for API testing
    # if current_user.role != 'admin':
    #     return {'success': False, 'message': 'Không có quyền'}, 403
    
    try:
        data = request.get_json()
        registration_ids = data.get('registration_ids', [])
        
        if not registration_ids:
            return jsonify({'success': False, 'message': 'Không có đăng ký nào được chọn'}), 400
        
        confirmed_count = 0
        for reg_id in registration_ids:
            registration = MealRegistration.query.get(reg_id)
            if registration and registration.menu and registration.menu.is_special:
                registration.is_confirmed = True
                registration.confirmed_at = datetime.utcnow()
                # registration.confirmed_by = current_user.id  # Skip for testing
                registration.confirmed_by = 1  # Use admin ID for testing
                confirmed_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Đã xác nhận {confirmed_count} suất cơm cải thiện'
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error in bulk_confirm_meal_registrations: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/meals/list')
@login_required
def list_meals():
    """Get all meals"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    from models import Menu
    menus = Menu.query.filter_by(is_active=True).all()
    
    return {
        'success': True,
        'menus': [{
            'id': m.id,
            'date': m.date.strftime('%Y-%m-%d'),
            'meal_type': m.meal_type,
            'dish_name': m.dish_name,
            'description': m.description,
            'image_url': m.image_url,
            'is_special': m.is_special,
            'is_vegetarian': m.is_vegetarian
        } for m in menus]
    }

@app.route('/admin/meals/add', methods=['POST'])
@login_required
def add_meal():
    """Add new meal to menu"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    from models import Menu
    from datetime import datetime
    
    data = request.json
    date_str = data.get('date')
    meal_type = data.get('meal_type')
    dish_name = data.get('dish_name', '').strip()
    description = data.get('description', '').strip()
    image_url = data.get('image_url', '').strip()
    is_special = data.get('is_special', False)
    is_vegetarian = data.get('is_vegetarian', False)
    
    if not date_str or not meal_type or not dish_name:
        return {'success': False, 'message': 'Vui lòng điền đầy đủ thông tin'}, 400
    
    try:
        date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return {'success': False, 'message': 'Ngày không hợp lệ'}, 400
    
    menu = Menu(
        date=date,
        meal_type=meal_type,
        dish_name=dish_name,
        description=description,
        image_url=image_url if image_url else None,
        is_special=is_special,
        is_vegetarian=is_vegetarian
    )
    
    db.session.add(menu)
    db.session.commit()
    
    return {'success': True, 'message': f'Đã thêm món {dish_name}'}

@app.route('/admin/meals/<int:meal_id>/edit', methods=['POST'])
@login_required
def edit_meal(meal_id):
    """Edit meal"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    from models import Menu
    from datetime import datetime
    
    menu = Menu.query.get_or_404(meal_id)
    data = request.json
    
    date_str = data.get('date')
    if date_str:
        try:
            menu.date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            return {'success': False, 'message': 'Ngày không hợp lệ'}, 400
    
    menu.meal_type = data.get('meal_type', menu.meal_type)
    menu.dish_name = data.get('dish_name', menu.dish_name).strip()
    menu.description = data.get('description', menu.description).strip()
    menu.is_special = data.get('is_special', menu.is_special)
    menu.is_vegetarian = data.get('is_vegetarian', menu.is_vegetarian)
    
    # Update image_url if provided
    if 'image_url' in data:
        menu.image_url = data['image_url'] if data['image_url'] else None
    
    db.session.commit()
    
    return {'success': True, 'message': f'Đã cập nhật món {menu.dish_name}'}

@app.route('/admin/meals/upload-image', methods=['POST'])
@login_required
def upload_meal_image():
    """Upload meal image"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    import os
    from werkzeug.utils import secure_filename
    
    if 'image' not in request.files:
        return {'success': False, 'message': 'Không có file ảnh'}, 400
    
    file = request.files['image']
    
    if file.filename == '':
        return {'success': False, 'message': 'Không có file được chọn'}, 400
    
    # Check file extension
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
    filename = secure_filename(file.filename)
    file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
    
    if file_ext not in allowed_extensions:
        return {'success': False, 'message': 'Định dạng file không hợp lệ'}, 400
    
    # Create upload directory if not exists
    upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'meals')
    os.makedirs(upload_dir, exist_ok=True)
    
    # Generate unique filename
    from datetime import datetime
    timestamp = now().strftime('%Y%m%d_%H%M%S')
    new_filename = f"{timestamp}_{filename}"
    file_path = os.path.join(upload_dir, new_filename)
    
    # Save file
    try:
        file.save(file_path)
        image_url = f"/static/uploads/meals/{new_filename}"
        return {'success': True, 'image_url': image_url, 'message': 'Tải ảnh lên thành công'}
    except Exception as e:
        return {'success': False, 'message': f'Lỗi khi lưu file: {str(e)}'}, 500

@app.route('/admin/meals/<int:meal_id>/delete', methods=['POST'])
@login_required
def delete_meal(meal_id):
    """Delete meal"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    from models import Menu
    menu = Menu.query.get_or_404(meal_id)
    
    dish_name = menu.dish_name
    db.session.delete(menu)
    db.session.commit()
    
    return {'success': True, 'message': f'Đã xóa món {dish_name}'}


@app.route('/admin/meals/template')
@login_required
def download_meal_template():
    """Download Excel template for meal import"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403

    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file
    from datetime import datetime, timedelta

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Thực đơn'

    # Header style
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='1976D2')
    center = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # New headers - simplified format
    headers = [
        'Ngày (YYYY-MM-DD)', 
        'Tên món (Thường)', 
        'Mô tả (Thường)', 
        'Tên món (Cải thiện)', 
        'Mô tả (Cải thiện)'
    ]
    col_widths = [20, 25, 30, 25, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    # Sample rows for single day import with auto-filled dates
    today_date = today()
    samples = [
        [(today_date + timedelta(days=0)).strftime('%Y-%m-%d'), 'Cơm gà xối mỡ', 'Gà chiên giòn, cơm trắng', 'Bún bò Huế', 'Bún bò đặc biệt'],
        [(today_date + timedelta(days=1)).strftime('%Y-%m-%d'), 'Cơm sườn nướng', 'Sườn nướng BBQ', 'Phở bò', 'Phở bò tái chín'],
        [(today_date + timedelta(days=2)).strftime('%Y-%m-%d'), 'Cơm chiên dương châu', 'Cơm chiên hải sản', 'Mì quảng', 'Mì quảng tôm thịt'],
    ]
    sample_fill = PatternFill(fill_type='solid', fgColor='F5F5F5')
    for r, row in enumerate(samples, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if c == 1:  # Date column - highlight with light blue
                cell.fill = PatternFill(fill_type='solid', fgColor='E3F2FD')
            elif r % 2 == 0:
                cell.fill = sample_fill


    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'template_thuc_don_{today.strftime("%Y%m%d")}.xlsx'
    )


@app.route('/admin/meals/import', methods=['POST'])
@login_required
def import_meals_excel():
    """Import meals from Excel file with preview and update support"""
    if current_user.role != 'admin':
        return {'success': False, 'message': 'Không có quyền'}, 403

    import openpyxl
    from datetime import datetime, timedelta
    from models import Menu

    # Check if this is a confirmation request
    if request.form.get('confirm_import'):
        return handle_import_confirmation()

    # Regular import with preview
    if 'file' not in request.files:
        return {'success': False, 'message': 'Không có file'}, 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return {'success': False, 'message': 'Chỉ hỗ trợ file .xlsx hoặc .xls'}, 400

    # Check if this is preview mode
    preview_only = request.form.get('preview_only', 'false').lower() == 'true'
    is_7_days = request.form.get('is_7_days', 'false').lower() == 'true'
    start_date_str = request.form.get('start_date')
    
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        errors = []
        preview_meals = []
        updates = 0

        # Parse start date for 7-day mode or use today as default
        start_date = None
        if is_7_days and start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                start_date = today()
        elif is_7_days:
            start_date = today()
        else:
            start_date = today()

        day_names = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6', 'Thứ 7', 'Chủ nhật']

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):
                continue

            # New format: Date, Normal_Name, Normal_Desc, Special_Name, Special_Desc
            date_val, normal_name, normal_desc, special_name, special_desc = (list(row) + [None]*5)[:5]

            # Convert None values to empty strings for validation
            normal_name = str(normal_name).strip() if normal_name is not None else ''
            special_name = str(special_name).strip() if special_name is not None else ''
            normal_desc = str(normal_desc).strip() if normal_desc is not None else ''
            special_desc = str(special_desc).strip() if special_desc is not None else ''

            # Validate required fields
            if not normal_name or not special_name:
                errors.append(f'Dòng {row_num}: Thiếu tên món (cả Thường và Cải thiện đều bắt buộc)')
                continue

            # Handle date logic
            date_obj = None
            
            if is_7_days:
                # For 7-day mode, use sequential dates from start_date
                day_offset = (row_num - 2) % 7  # Cycle through 7 days
                date_obj = start_date + timedelta(days=day_offset)
            else:
                # For single day mode:
                # - Empty date => SKIP row (per new rule)
                # - Invalid overflow date (e.g. 2026-04-34) => SKIP row
                if not date_val or str(date_val).strip().upper() in ['', 'IGNORE', 'AUTO', 'NULL']:
                    errors.append(f'Dòng {row_num}: Ngày trống/không hợp lệ - đã bỏ qua dòng')
                    continue
                else:
                    if isinstance(date_val, str):
                        date_str = date_val.strip()
                        try:
                            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                        except ValueError:
                            errors.append(
                                f'Dòng {row_num}: Ngày "{date_str}" không hợp lệ (vượt ngày trong tháng hoặc sai định dạng) - đã bỏ qua dòng'
                            )
                            continue
                    elif hasattr(date_val, 'date'):
                        date_obj = date_val.date()
                    else:
                        errors.append(f'Dòng {row_num}: Giá trị ngày không hợp lệ - đã bỏ qua dòng')
                        continue

            # Strict single-month rule:
            # Determine target month from first valid date in file.
            if 'target_month' not in locals():
                target_month = (date_obj.year, date_obj.month)
            else:
                if (date_obj.year, date_obj.month) != target_month:
                    errors.append(
                        f'Dòng {row_num}: Ngày {date_obj.strftime("%Y-%m-%d")} khác tháng mục tiêu {target_month[0]}-{target_month[1]:02d} - đã bỏ qua dòng'
                    )
                    continue

            # Check for existing meals on this date
            existing_normal = Menu.query.filter_by(
                date=date_obj,
                meal_type='lunch',
                is_special=False,
                is_vegetarian=True
            ).first()
            
            existing_special = Menu.query.filter_by(
                date=date_obj,
                meal_type='lunch',
                is_special=True,
                is_vegetarian=False
            ).first()

            will_update = existing_normal or existing_special
            if will_update:
                updates += 1

            # Create preview data (ensure all values are JSON serializable)
            preview_meal = {
                'row_num': int(row_num),
                'date': date_obj.strftime('%Y-%m-%d'),
                'day_name': day_names[date_obj.weekday()],
                'normal_name': normal_name[:200],  # Already cleaned above
                'normal_desc': normal_desc,        # Already cleaned above
                'special_name': special_name[:200], # Already cleaned above
                'special_desc': special_desc,       # Already cleaned above
                'will_update': bool(will_update),
                'has_error': False,
                'existing_normal_id': int(existing_normal.id) if existing_normal else None,
                'existing_special_id': int(existing_special.id) if existing_special else None
            }
            
            preview_meals.append(preview_meal)

        # If preview mode, return preview data (ensure all data is JSON serializable)
        if preview_only:
            return {
                'success': True,
                'preview_data': {
                    'total_rows': int(len(preview_meals)),
                    'total_meals': int(len(preview_meals) * 2),
                    'updates': int(updates),
                    'errors': list(errors),
                    'meals': list(preview_meals),
                    'date_range': {
                        'start': str(min(m['date'] for m in preview_meals)) if preview_meals else '',
                        'end': str(max(m['date'] for m in preview_meals)) if preview_meals else ''
                    }
                }
            }

    except Exception as e:
        return {'success': False, 'message': f'Lỗi xử lý file: {str(e)}'}, 500


def handle_import_confirmation():
    """Handle the actual import after user confirmation"""
    import json
    from datetime import datetime
    from models import Menu
    
    try:
        # Get preview data with proper error handling
        preview_data_str = request.form.get('preview_data')
        # Debug: Received preview_data_str
        
        if not preview_data_str:
            # Debug: No preview_data in request.form
            return {'success': False, 'message': 'Không có dữ liệu preview'}, 400
        
        try:
            preview_data = json.loads(preview_data_str)
            # Debug: Parsed preview_data keys
        except (json.JSONDecodeError, TypeError) as e:
            # Debug: JSON decode error
            return {'success': False, 'message': f'Lỗi phân tích dữ liệu preview: {str(e)}'}, 400
        
        if not preview_data or not isinstance(preview_data, dict):
            # Debug: Invalid preview_data type
            return {'success': False, 'message': 'Dữ liệu preview không hợp lệ'}, 400
        
        is_7_days = request.form.get('is_7_days', 'false').lower() == 'true'
        # Debug: is_7_days value
        
        added, updated = 0, 0
        
        meals_data = preview_data.get('meals', [])
        # Debug: Found meals in preview_data
        
        if not meals_data:
            return {'success': False, 'message': 'Không có dữ liệu món ăn để import'}, 400
        
        for meal_data in meals_data:
            if not meal_data or meal_data.get('has_error'):
                continue
                
            # Validate required fields
            if not meal_data.get('date') or not meal_data.get('normal_name') or not meal_data.get('special_name'):
                continue
                
            date_obj = datetime.strptime(meal_data['date'], '%Y-%m-%d').date()
            
            # Handle Normal meal
            if meal_data.get('existing_normal_id'):
                # Update existing normal meal
                existing_normal = Menu.query.get(meal_data['existing_normal_id'])
                if existing_normal:
                    existing_normal.dish_name = meal_data['normal_name']
                    existing_normal.description = meal_data.get('normal_desc') or None
                    updated += 1
            else:
                # Create new normal meal
                normal_menu = Menu(
                    date=date_obj,
                    meal_type='lunch',
                    dish_name=meal_data['normal_name'],
                    description=meal_data.get('normal_desc') or None,
                    is_special=False,
                    is_vegetarian=True,
                    is_active=True
                )
                db.session.add(normal_menu)
                added += 1
            
            # Handle Special meal
            if meal_data.get('existing_special_id'):
                # Update existing special meal
                existing_special = Menu.query.get(meal_data['existing_special_id'])
                if existing_special:
                    existing_special.dish_name = meal_data['special_name']
                    existing_special.description = meal_data.get('special_desc') or None
                    updated += 1
            else:
                # Create new special meal
                special_menu = Menu(
                    date=date_obj,
                    meal_type='lunch',
                    dish_name=meal_data['special_name'],
                    description=meal_data.get('special_desc') or None,
                    is_special=True,
                    is_vegetarian=False,
                    is_active=True
                )
                db.session.add(special_menu)
                added += 1

        db.session.commit()

        # Create success message
        msg_parts = []
        if added > 0:
            msg_parts.append(f'Đã thêm {added} món mới')
        if updated > 0:
            msg_parts.append(f'Đã cập nhật {updated} món')
            
        message = ', '.join(msg_parts) if msg_parts else 'Import hoàn tất'

        return {
            'success': True,
            'message': str(message),
            'added': int(added),
            'updated': int(updated),
            'total': int(added + updated)
        }

    except Exception as e:
        db.session.rollback()
        return {'success': False, 'message': f'Lỗi import: {str(e)}'}, 500


# ============= END MEAL MANAGEMENT ROUTES =============

@app.route('/overtime', methods=['GET', 'POST'])
@login_required
def overtime():
    # Check if user has permission to register overtime
    # Managers (can_approve=True) automatically have registration permission
    if not current_user.can_register and not current_user.can_approve:
        flash('Bạn không có quyền đăng ký tăng ca', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        from datetime import datetime, time
        
        # Get form data
        selected_date = request.form.get('selected_date')
        overtime_hours = request.form.get('overtime_hours')  # This is number of people
        reason = request.form.get('reason', '')
        
        if not all([selected_date, overtime_hours]):
            flash('Vui lòng điền đầy đủ thông tin', 'error')
            return redirect(url_for('overtime'))
        
        # Parse date and number of people
        try:
            overtime_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
            num_people = int(overtime_hours)
            
            if num_people <= 0:
                flash('Số người phải lớn hơn 0', 'error')
                return redirect(url_for('overtime'))
                
        except ValueError:
            flash('Dữ liệu không hợp lệ', 'error')
            return redirect(url_for('overtime'))
        
        # Không cho phép đăng ký tăng ca cho ngày hiện tại trở về trước hoặc ngày hiện tại sau 16h
        current_time = now()
        today_date = current_time.date()
        current_hour = current_time.hour
        
        if overtime_date < today_date:
            flash('Không thể đăng ký tăng ca cho ngày đã qua', 'error')
            return redirect(url_for('overtime'))
        elif overtime_date == today_date and current_hour >= 16:
            flash('Không thể đăng ký tăng ca cho hôm nay sau 16h', 'error')
            return redirect(url_for('overtime'))

        # Get department name from relationship
        department_name = current_user.dept.name if current_user.dept else 'Chưa xác định'
        
        # Check if already registered for this date and department
        # 1 ngày chỉ có 1 lần đăng ký per department - update if exists
        existing_request = OvertimeRequest.query.filter_by(
            overtime_date=overtime_date,
            department=department_name
        ).first()
        
        if existing_request:
            # Update existing request
            existing_request.user_id = current_user.id
            existing_request.employee_id = current_user.employee_id
            existing_request.employee_name = current_user.name
            existing_request.total_hours = num_people  # Store number of people in total_hours field
            existing_request.reason = reason
            existing_request.status = 'pending'
            existing_request.updated_at = datetime.utcnow()
            
            db.session.commit()
            flash(f'Đã cập nhật đăng ký tăng ca cho ngày {overtime_date.strftime("%d/%m/%Y")}', 'success')
            return redirect(url_for('overtime'))
        else:
            # Create new overtime request
            # Use dummy time values since we only care about number of people
            overtime_request = OvertimeRequest(
                user_id=current_user.id,
                employee_id=current_user.employee_id,
                employee_name=current_user.name,
                department=department_name,
                overtime_date=overtime_date,
                start_time=time(17, 0),  # Dummy value
                end_time=time(19, 0),    # Dummy value
                total_hours=num_people,  # Store number of people here
                reason=reason,
                status='pending'
            )
            
            db.session.add(overtime_request)
            db.session.commit()
            
            flash(f'Đã gửi yêu cầu tăng ca cho ngày {overtime_date.strftime("%d/%m/%Y")}', 'success')
            return redirect(url_for('overtime'))
    
    # GET request - show form
    return render_template('overtime.html')

@app.route('/overtime/my-requests')
@login_required
def get_my_overtime_requests():
    """Get current user's overtime requests"""
    try:
        requests = OvertimeRequest.query.filter_by(user_id=current_user.id).order_by(OvertimeRequest.created_at.desc()).all()
        
        result_list = []
        for r in requests:
            # Get approver info if request was approved/rejected
            approver_name = ''
            approver_employee_id = ''
            if r.manager_id:
                try:
                    approver = User.query.get(r.manager_id)
                    if approver:
                        approver_name = approver.name or ''
                        approver_employee_id = approver.employee_id or ''
                except Exception as e:
                    print(f"Error getting approver {r.manager_id}: {e}")
            
            result_list.append({
                'id': r.id,
                'overtime_date': r.overtime_date.strftime('%Y-%m-%d'),
                'start_time': r.start_time.strftime('%H:%M') if r.start_time else '00:00',
                'end_time': r.end_time.strftime('%H:%M') if r.end_time else '00:00',
                'total_hours': float(r.total_hours) if r.total_hours else 0,
                'reason': r.reason or '',
                'status': r.status,
                'manager_comment': r.manager_comment or '',
                'created_at': r.created_at.strftime('%Y-%m-%d %H:%M') if r.created_at else '',
                # Add approver info
                'approver_name': approver_name,
                'approver_username': approver_employee_id
            })
        
        return jsonify({
            'success': True,
            'requests': result_list
        })
    except Exception as e:
        print(f"Error in get_my_overtime_requests: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/overtime/<int:request_id>/cancel', methods=['POST'])
@login_required
def cancel_overtime_request(request_id):
    """Cancel overtime request (only if pending)"""
    overtime_request = OvertimeRequest.query.get_or_404(request_id)
    
    # Check if user owns this request
    if overtime_request.user_id != current_user.id:
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    # Can only cancel pending requests
    if overtime_request.status != 'pending':
        return {'success': False, 'message': 'Chỉ có thể hủy yêu cầu đang chờ duyệt'}, 400
    
    db.session.delete(overtime_request)
    db.session.commit()
    
    return {'success': True, 'message': 'Đã hủy yêu cầu tăng ca'}

@app.route('/overtime/<int:request_id>/self-approve', methods=['POST'])
@login_required
def self_approve_overtime_request(request_id):
    """Self-approve overtime request (only for highest level users)"""
    overtime_request = OvertimeRequest.query.get_or_404(request_id)
    
    # Check if user owns this request
    if overtime_request.user_id != current_user.id:
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    # Can only approve pending requests
    if overtime_request.status != 'pending':
        return {'success': False, 'message': 'Yêu cầu này đã được xử lý'}, 400
    
    # Check if user has approval permission
    if not current_user.can_approve:
        return {'success': False, 'message': 'Bạn không có quyền tự duyệt'}, 403
    
    # Approve the request
    overtime_request.status = 'approved'
    overtime_request.manager_id = current_user.id
    overtime_request.manager_approved_at = datetime.utcnow()
    overtime_request.manager_comment = 'Tự duyệt (cấp cao nhất trong phòng)'
    
    db.session.commit()
    
    return {'success': True, 'message': 'Đã duyệt yêu cầu tăng ca'}

# ============= LEAVE REQUEST ROUTES =============
@app.route('/leave', methods=['GET', 'POST'])
@login_required
def leave():
    # Check if user has permission to register leave
    # Managers (can_approve=True) automatically have registration permission
    if not current_user.can_register and not current_user.can_approve:
        flash('Bạn không có quyền đăng ký nghỉ phép', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        from datetime import datetime
        
        # Get form data
        leave_type = request.form.get('leave_type')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        total_days = request.form.get('total_days')
        reason = request.form.get('reason')
        emergency_contact = request.form.get('emergency_contact')
        emergency_phone = request.form.get('emergency_phone')
        
        if not all([leave_type, start_date, end_date, total_days, reason]):
            flash('Vui lòng điền đầy đủ thông tin', 'error')
            return redirect(url_for('leave'))
        
        # Parse dates
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            total_days_int = int(total_days)
        except ValueError:
            flash('Định dạng ngày không hợp lệ', 'error')
            return redirect(url_for('leave'))
        
        # Validate dates
        if start_date_obj > end_date_obj:
            flash('Ngày bắt đầu phải trước ngày kết thúc', 'error')
            return redirect(url_for('leave'))
        
        # Create leave request
        leave_request = LeaveRequest(
            user_id=current_user.id,
            employee_id=current_user.employee_id,
            employee_name=current_user.name,
            department=current_user.dept.name if current_user.dept else 'Chưa xác định',
            leave_type=leave_type,
            start_date=start_date_obj,
            end_date=end_date_obj,
            total_days=total_days_int,
            reason=reason,
            emergency_contact=emergency_contact if leave_type == 'sick' else None,
            emergency_phone=emergency_phone if leave_type == 'sick' else None,
            status='pending'
        )
        
        db.session.add(leave_request)
        db.session.commit()
        
        flash('Đã gửi yêu cầu nghỉ phép thành công!', 'success')
        return redirect(url_for('leave'))
    
    # GET request - show form
    # Check if user is manager for sidebar
    is_manager = current_user.can_approve or current_user.role == 'admin'
    return render_template('leave.html', is_manager=is_manager)


@app.route('/leave/my-requests')
@login_required
def get_my_leave_requests():
    """Get current user's leave requests"""
    requests = LeaveRequest.query.filter_by(user_id=current_user.id).order_by(LeaveRequest.created_at.desc()).all()
    
    return {
        'success': True,
        'requests': [{
            'id': r.id,
            'leave_type': r.leave_type,
            'start_date': r.start_date.strftime('%Y-%m-%d'),
            'end_date': r.end_date.strftime('%Y-%m-%d'),
            'total_days': r.total_days,
            'reason': r.reason,
            'status': r.status,
            'manager_comment': r.manager_comment,
            'emergency_contact': r.emergency_contact,
            'emergency_phone': r.emergency_phone,
            'created_at': r.created_at.strftime('%Y-%m-%d %H:%M')
        } for r in requests]
    }


@app.route('/leave/<int:request_id>/cancel', methods=['POST'])
@login_required
def cancel_leave_request(request_id):
    """Cancel leave request (only if pending)"""
    leave_request = LeaveRequest.query.get_or_404(request_id)
    
    # Check if user owns this request
    if leave_request.user_id != current_user.id:
        return {'success': False, 'message': 'Không có quyền'}, 403
    
    # Can only cancel pending requests
    if leave_request.status != 'pending':
        return {'success': False, 'message': 'Chỉ có thể hủy yêu cầu đang chờ duyệt'}, 400
    
    db.session.delete(leave_request)
    db.session.commit()
    
    return {'success': True, 'message': 'Đã hủy yêu cầu nghỉ phép'}

@app.route('/exit-entry', methods=['GET', 'POST'])
@login_required
def exit_entry():
    """Exit/Entry request form - similar to leave request"""
    if request.method == 'POST':
        from datetime import datetime, date
        
        # Get form data
        request_date = request.form.get('request_date')
        exit_time = request.form.get('exit_time')
        entry_time = request.form.get('entry_time')
        reason = request.form.get('reason')
        
        # Validation
        if not request_date or not reason:
            flash('Vui lòng điền đầy đủ thông tin bắt buộc', 'error')
            return redirect(url_for('exit_entry'))
        
        # Convert date string to date object
        try:
            request_date_obj = datetime.strptime(request_date, '%Y-%m-%d').date()
        except ValueError:
            flash('Ngày không hợp lệ', 'error')
            return redirect(url_for('exit_entry'))
        
        # Convert time strings to time objects
        exit_time_obj = None
        entry_time_obj = None
        
        if exit_time:
            try:
                exit_time_obj = datetime.strptime(exit_time, '%H:%M').time()
            except ValueError:
                flash('Thời gian ra không hợp lệ', 'error')
                return redirect(url_for('exit_entry'))
        
        if entry_time:
            try:
                entry_time_obj = datetime.strptime(entry_time, '%H:%M').time()
            except ValueError:
                flash('Thời gian vào không hợp lệ', 'error')
                return redirect(url_for('exit_entry'))
        
        # Create exit/entry request
        exit_entry_request = ExitEntryRequest(
            user_id=current_user.id,
            employee_id=current_user.employee_id,
            employee_name=current_user.name,
            department=current_user.dept.name if current_user.dept else 'Chưa xác định',
            request_date=request_date_obj,
            exit_time=exit_time_obj,
            entry_time=entry_time_obj,
            reason=reason,
            status='pending'
        )
        
        db.session.add(exit_entry_request)
        db.session.commit()
        
        flash('Đã gửi đơn xin ra vào công ty thành công!', 'success')
        return redirect(url_for('exit_entry'))
    
    # GET request - show form
    is_manager = current_user.can_approve or current_user.role == 'admin'
    
    return render_template('exit_entry.html', is_manager=is_manager)

@app.route('/exit-entry/my-requests')
@login_required
def my_exit_entry_requests():
    """Get user's exit/entry requests as JSON"""
    requests = ExitEntryRequest.query.filter_by(user_id=current_user.id).order_by(ExitEntryRequest.created_at.desc()).all()
    
    return {
        'success': True,
        'requests': [{
            'id': r.id,
            'request_date': r.request_date.strftime('%Y-%m-%d'),
            'exit_time': r.exit_time.strftime('%H:%M') if r.exit_time else None,
            'entry_time': r.entry_time.strftime('%H:%M') if r.entry_time else None,
            'reason': r.reason,
            'status': r.status,
            'manager_comment': r.manager_comment,
            'created_at': r.created_at.strftime('%Y-%m-%d %H:%M')
        } for r in requests]
    }

@app.route('/exit-entry/<int:request_id>/cancel', methods=['POST'])
@login_required
def cancel_exit_entry_request(request_id):
    """Cancel exit/entry request"""
    exit_entry_request = ExitEntryRequest.query.get_or_404(request_id)
    
    # Check if user owns this request
    if exit_entry_request.user_id != current_user.id:
        return {'success': False, 'message': 'Không có quyền hủy yêu cầu này'}, 403
    
    # Check if request can be cancelled
    if exit_entry_request.status != 'pending':
        return {'success': False, 'message': 'Chỉ có thể hủy yêu cầu đang chờ duyệt'}, 400
    
    # Delete the request
    db.session.delete(exit_entry_request)
    db.session.commit()
    
    return {'success': True, 'message': 'Đã hủy yêu cầu thành công'}

@app.route('/meals', methods=['GET', 'POST'])
@login_required
def meals():
    from datetime import datetime, timedelta
    
    if request.method == 'POST':
        date = request.form.get('date')
        meal_id = request.form.get('meal_id')
        
        if not date or not meal_id:
            flash('Vui lòng chọn món ăn', 'error')
            return redirect(url_for('meals'))
        
        # Check if registration already exists
        existing = MealRegistration.query.filter_by(
            user_id=current_user.id,
            date=date
        ).first()
        
        if existing:
            # Update existing registration
            existing.meal_id = meal_id
            existing.updated_at = datetime.utcnow()
        else:
            # Create new registration
            meal_registration = MealRegistration(
                user_id=current_user.id,
                date=date,
                meal_id=meal_id,
                has_meal=True
            )
            db.session.add(meal_registration)
        
        db.session.commit()
        flash('Đăng ký suất ăn thành công!', 'success')
        return redirect(url_for('meals'))
    
    # Get current week (Monday to Sunday)
    today_date = today()
    start_of_week = today_date - timedelta(days=today_date.weekday())  # Monday
    
    days_of_week = []
    day_names = ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN']
    
    for i in range(7):
        day = start_of_week + timedelta(days=i)
        days_of_week.append({
            'date': day.strftime('%Y-%m-%d'),
            'day_name': day_names[i],
            'day_num': day.day
        })
    
    # Get user's meal registrations
    registrations = MealRegistration.query.filter_by(user_id=current_user.id).order_by(MealRegistration.date.desc()).all()
    
    return render_template('meals.html', 
                         days_of_week=days_of_week,
                         registrations=registrations)

@app.route('/meals/menu/<date>')
@login_required
def get_menu_for_date(date):
    """Get menu for specific date"""
    from models import Menu
    from datetime import datetime
    
    try:
        date_obj = datetime.strptime(date, '%Y-%m-%d').date()
    except ValueError:
        return {'success': False, 'message': 'Ngày không hợp lệ'}, 400
    
    menus = Menu.query.filter_by(date=date_obj, is_active=True).all()
    
    # Get user's registration for this date
    registration = MealRegistration.query.filter_by(
        user_id=current_user.id,
        date=date_obj
    ).first()
    
    # Check if registration is locked (after 4 PM for next day)
    current_time = now()
    today_date = current_time.date()
    current_hour = current_time.hour
    
    # If it's after 4 PM (16:00) and the selected date is tomorrow, lock registration
    is_locked = False
    if current_hour >= 16:
        from datetime import timedelta
        tomorrow = today_date + timedelta(days=1)
        if date_obj == tomorrow:
            is_locked = True
    
    # Also lock if the date is in the past
    if date_obj < today_date:
        is_locked = True
    
    return {
        'success': True,
        'menus': [{
            'id': m.id,
            'dish_name': m.dish_name,
            'description': m.description,
            'image_url': m.image_url,
            'meal_type': m.meal_type,
            'is_special': m.is_special,
            'is_vegetarian': m.is_vegetarian
        } for m in menus],
        'selected_meal_id': registration.meal_id if registration else None,
        'is_locked': is_locked
    }

@app.route('/meals/week-registrations')
@login_required
def get_week_registrations():
    """Get user's meal registrations for the current week"""
    from datetime import datetime, timedelta
    from models import Menu
    
    # Get week offset from query params
    week_offset = int(request.args.get('offset', 0))
    
    # Calculate week start (Monday)
    today_date = today()
    today_with_offset = today_date + timedelta(days=week_offset * 7)
    week_start = today_with_offset - timedelta(days=today_with_offset.weekday())
    
    # Get registrations for the week
    registrations = {}
    for i in range(7):
        day = week_start + timedelta(days=i)
        day_str = day.strftime('%Y-%m-%d')
        
        reg = MealRegistration.query.filter_by(
            user_id=current_user.id,
            date=day
        ).first()
        
        if reg and reg.menu:
            registrations[day_str] = {
                'has_registration': True,
                'is_special': reg.menu.is_special
            }
        else:
            registrations[day_str] = {
                'has_registration': False,
                'is_special': False
            }
    
    return {'success': True, 'registrations': registrations}

@app.route('/meals/register', methods=['POST'])
@login_required
def register_meal():
    """Register meal for user"""
    from datetime import datetime, timedelta
    
    data = request.json
    date_str = data.get('date')
    meal_id = data.get('meal_id')
    
    if not date_str or not meal_id:
        return {'success': False, 'message': 'Thiếu thông tin'}, 400
    
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return {'success': False, 'message': 'Ngày không hợp lệ'}, 400
    
    # Check if registration is locked
    current_time = now()
    today_date = current_time.date()
    current_hour = current_time.hour
    
    # If it's after 4 PM (16:00) and trying to register for tomorrow, deny
    if current_hour >= 16:
        tomorrow = today_date + timedelta(days=1)
        if date_obj == tomorrow:
            return {'success': False, 'message': 'Đã quá thời gian đăng ký cho ngày mai (sau 16h)'}, 403
    
    # Don't allow registration for past dates
    if date_obj < today_date:
        return {'success': False, 'message': 'Không thể đăng ký cho ngày đã qua'}, 403
    
    # Check if registration exists
    existing = MealRegistration.query.filter_by(
        user_id=current_user.id,
        date=date_obj
    ).first()
    
    if existing:
        existing.meal_id = meal_id
        existing.has_meal = True
        existing.updated_at = datetime.utcnow()
    else:
        registration = MealRegistration(
            user_id=current_user.id,
            date=date_obj,
            meal_id=meal_id,
            has_meal=True
        )
        db.session.add(registration)
    
    db.session.commit()
    
    return {'success': True, 'message': 'Đăng ký thành công!'}

@app.route('/meals/cancel', methods=['POST'])
@login_required
def cancel_meal():
    """Cancel meal registration"""
    from datetime import datetime, timedelta
    
    data = request.json
    date_str = data.get('date')
    
    if not date_str:
        return {'success': False, 'message': 'Thiếu thông tin'}, 400
    
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return {'success': False, 'message': 'Ngày không hợp lệ'}, 400
    
    # Check if cancellation is locked
    current_time = now()
    today_date = current_time.date()
    current_hour = current_time.hour
    
    # If it's after 4 PM (16:00) and trying to cancel for tomorrow, deny
    if current_hour >= 16:
        tomorrow = today_date + timedelta(days=1)
        if date_obj == tomorrow:
            return {'success': False, 'message': 'Đã quá thời gian hủy đăng ký cho ngày mai (sau 16h)'}, 403
    
    # Don't allow cancellation for past dates
    if date_obj < today_date:
        return {'success': False, 'message': 'Không thể hủy đăng ký cho ngày đã qua'}, 403
    
    # Find and delete registration
    existing = MealRegistration.query.filter_by(
        user_id=current_user.id,
        date=date_obj
    ).first()
    
    if existing:
        db.session.delete(existing)
        db.session.commit()
        return {'success': True, 'message': 'Đã hủy đăng ký!'}
    else:
        return {'success': False, 'message': 'Không tìm thấy đăng ký'}, 404

@app.route('/profile')
@login_required
def profile():
    from datetime import datetime, timedelta
    from sqlalchemy import func
    
    # Calculate total overtime hours for current year
    current_year = now().year
    total_overtime_hours = db.session.query(
        func.sum(OvertimeRequest.total_hours)
    ).filter(
        OvertimeRequest.user_id == current_user.id,
        OvertimeRequest.status == 'approved',
        func.year(OvertimeRequest.overtime_date) == current_year
    ).scalar() or 0
    
    # Count total meals registered this year
    total_meals = MealRegistration.query.filter(
        MealRegistration.user_id == current_user.id,
        func.year(MealRegistration.date) == current_year
    ).count()
    
    # Check if today has special meal registration
    today_date = today()
    today_meal = MealRegistration.query.filter(
        MealRegistration.user_id == current_user.id,
        MealRegistration.date == today_date
    ).first()
    
    has_special_meal_today = False
    if today_meal and today_meal.menu:
        has_special_meal_today = today_meal.menu.is_special
    
    return render_template('profile.html', 
                         user=current_user,
                         total_overtime_hours=float(total_overtime_hours),
                         total_meals=total_meals,
                         has_special_meal_today=has_special_meal_today,
                         now=now(),
                         timedelta=timedelta)

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    form = ChangePasswordForm()
    
    if form.validate_on_submit():
        # Verify current password
        if not check_password_hash(current_user.password, form.current_password.data):
            flash('Current password is incorrect', 'error')
            return render_template('change_password.html', form=form)
        
        # Update password
        current_user.password = generate_password_hash(form.new_password.data)
        db.session.commit()
        flash('Password changed successfully!', 'success')
        return redirect(url_for('profile'))
    
    return render_template('change_password.html', form=form)

@app.route('/logout')
def logout():
    """Logout user and clear all session data"""
    # Force remove user from session
    session.pop('_user_id', None)
    session.pop('user_name', None)
    session.pop('_fresh', None)
    
    # Logout user
    if current_user.is_authenticated:
        logout_user()
    
    # Clear all session data
    session.clear()
    
    # Create response with redirect
    response = redirect(url_for('login'))
    
    # Delete all cookies
    response.set_cookie('remember_token', '', expires=0, path='/')
    response.set_cookie('session', '', expires=0, path='/')
    
    return response

# PWA Routes
@app.route('/static/manifest.json')
def manifest():
    return app.send_static_file('manifest.json')

@app.route('/static/service-worker.js')
def service_worker():
    return app.send_static_file('service-worker.js')

# ============= AUTO REGISTER MEALS SCHEDULER =============
def auto_register_meals_for_user(user_id, days=30):
    """
    Tự động đăng ký suất ăn cho 1 user cụ thể
    Dùng khi tạo user mới hoặc kích hoạt lại user
    
    Args:
        user_id: ID của user cần đăng ký
        days: Số ngày cần đăng ký (mặc định 30)
    
    Returns:
        dict: Kết quả đăng ký {success: bool, registered: int, skipped: int, message: str}
    """
    try:
        from models import Menu, MealRegistration
        from sqlalchemy.exc import IntegrityError
        
        user = User.query.get(user_id)
        if not user:
            return {'success': False, 'registered': 0, 'skipped': 0, 'message': 'User không tồn tại'}
        
        # Kiểm tra điều kiện - chỉ cần work_status = 'working'
        if user.work_status != 'working':
            return {'success': False, 'registered': 0, 'skipped': 0, 'message': 'User không đang làm việc'}
        
        # Kiểm tra mã nhân viên
        if not (user.employee_id.startswith('10') or user.employee_id.startswith('20')):
            return {'success': False, 'registered': 0, 'skipped': 0, 'message': 'Mã nhân viên không thuộc 10xx hoặc 20xx'}
        
        today_date = today()
        registered_count = 0
        skipped_count = 0
        
        print(f"[AUTO REGISTER USER] Bắt đầu đăng ký cho {user.name} ({user.employee_id}) - {days} ngày")
        
        # Lặp qua số ngày cần đăng ký
        for day_offset in range(1, days + 1):
            target_date = today_date + timedelta(days=day_offset)
            
            # Kiểm tra đã có đăng ký chưa - QUAN TRỌNG: Chỉ 1 đăng ký/user/ngày
            existing = MealRegistration.query.filter_by(
                user_id=user.id,
                date=target_date
            ).first()
            
            if existing:
                skipped_count += 1
                continue
            
            # Lấy menu bình thường
            normal_menu = Menu.query.filter(
                Menu.date == target_date,
                Menu.is_special == False,
                Menu.is_active == True,
                Menu.meal_type == 'lunch'
            ).first()
            
            if not normal_menu:
                skipped_count += 1
                continue
            
            # Tạo đăng ký mới với xử lý lỗi IntegrityError
            try:
                new_registration = MealRegistration(
                    user_id=user.id,
                    date=target_date,
                    meal_id=normal_menu.id,
                    meal_type='lunch',
                    has_meal=True,
                    notes='Tự động đăng ký bởi hệ thống'
                )
                db.session.add(new_registration)
                db.session.flush()  # Flush để kiểm tra constraint ngay lập tức
                registered_count += 1
                
            except IntegrityError:
                # Nếu có lỗi unique constraint, rollback và bỏ qua
                db.session.rollback()
                skipped_count += 1
                print(f"[AUTO REGISTER USER] Ngày {target_date}: Đã có đăng ký (IntegrityError)")
                continue
        
        db.session.commit()
        
        print(f"[AUTO REGISTER USER] Hoàn thành: {registered_count} đăng ký mới, {skipped_count} bỏ qua")
        
        return {
            'success': True,
            'registered': registered_count,
            'skipped': skipped_count,
            'message': f'Đã đăng ký {registered_count} ngày, bỏ qua {skipped_count} ngày'
        }
        
    except Exception as e:
        db.session.rollback()
        print(f"[AUTO REGISTER USER] Lỗi: {str(e)}")
        return {'success': False, 'registered': 0, 'skipped': 0, 'message': f'Lỗi: {str(e)}'}


def auto_register_meals_for_30_days():
    """
    Tự động đăng ký suất ăn bình thường cho 30 ngày tới
    Chỉ áp dụng cho nhân viên có mã bắt đầu bằng 10 hoặc 20
    Chạy tự động lúc 16h hàng ngày
    """
    with app.app_context():
        try:
            from datetime import timedelta
            from sqlalchemy import or_
            from sqlalchemy.exc import IntegrityError
            today_date = today()  # Use timezone-aware today
            start_date = today_date + timedelta(days=1)  # Bắt đầu từ ngày mai
            end_date = today_date + timedelta(days=30)   # Đến 30 ngày sau
            
            print(f"\n=== [AUTO REGISTER] Bắt đầu tự động đăng ký 30 ngày ({start_date.strftime('%d/%m/%Y')} → {end_date.strftime('%d/%m/%Y')}) ===")
            
            # Lấy tất cả user đang làm việc và có mã bắt đầu bằng 10 hoặc 20
            active_users = User.query.filter(
                User.work_status == 'working',
                or_(
                    User.employee_id.like('10%'),
                    User.employee_id.like('20%')
                )
            ).all()
            
            print(f"[AUTO REGISTER] Tìm thấy {len(active_users)} user đang làm việc (mã 10xx, 20xx)")
            
            from models import Menu, MealRegistration
            
            total_registered = 0
            total_skipped = 0
            total_errors = 0
            days_processed = 0
            
            # Lặp qua 30 ngày
            for day_offset in range(1, 31):
                target_date = today_date + timedelta(days=day_offset)
                
                # Lấy menu bình thường cho ngày này
                normal_menu = Menu.query.filter(
                    Menu.date == target_date,
                    Menu.is_special == False,
                    Menu.is_active == True,
                    Menu.meal_type == 'lunch'
                ).first()
                
                if not normal_menu:
                    print(f"[AUTO REGISTER] Ngày {target_date.strftime('%d/%m')}: Không có menu bình thường - Bỏ qua")
                    continue
                
                day_registered = 0
                day_skipped = 0
                day_errors = 0
                
                for user in active_users:
                    # Kiểm tra xem user đã đăng ký chưa (dù bình thường hay cải thiện)
                    existing_registration = MealRegistration.query.filter_by(
                        user_id=user.id,
                        date=target_date
                    ).first()
                    
                    if existing_registration:
                        day_skipped += 1
                    else:
                        # Tự động đăng ký suất ăn bình thường với xử lý lỗi
                        try:
                            new_registration = MealRegistration(
                                user_id=user.id,
                                date=target_date,
                                meal_id=normal_menu.id,
                                meal_type='lunch',
                                has_meal=True,
                                notes='Tự động đăng ký bởi hệ thống'
                            )
                            db.session.add(new_registration)
                            db.session.flush()  # Flush để kiểm tra constraint ngay
                            day_registered += 1
                            
                        except IntegrityError:
                            # Nếu có lỗi unique constraint, rollback và bỏ qua
                            db.session.rollback()
                            day_errors += 1
                            print(f"[AUTO REGISTER] Lỗi duplicate cho {user.employee_id} ngày {target_date.strftime('%d/%m')}")
                            continue
                
                # Commit sau mỗi ngày để tránh rollback toàn bộ
                try:
                    db.session.commit()
                except Exception as e:
                    print(f"[AUTO REGISTER] Lỗi commit ngày {target_date.strftime('%d/%m')}: {str(e)}")
                    db.session.rollback()
                    day_errors += day_registered
                    day_registered = 0
                
                total_registered += day_registered
                total_skipped += day_skipped
                total_errors += day_errors
                days_processed += 1
                
                print(f"[AUTO REGISTER] Ngày {target_date.strftime('%d/%m')}: {normal_menu.dish_name[:30]} | Đã có: {day_skipped} | Tự động: {day_registered} | Lỗi: {day_errors}")
            
            print(f"\n[AUTO REGISTER] ✓ Hoàn thành!")
            print(f"[AUTO REGISTER] Tổng: {days_processed} ngày | {len(active_users)} users | Đã có: {total_skipped} | Tự động: {total_registered} | Lỗi: {total_errors}")
            
        except Exception as e:
            print(f"[AUTO REGISTER] ❌ Lỗi: {str(e)}")
            db.session.rollback()

# Initialize scheduler
scheduler = BackgroundScheduler()
scheduler.add_job(
    func=auto_register_meals_for_30_days,
    trigger=CronTrigger(hour=16, minute=0, timezone='Asia/Ho_Chi_Minh'),  # 16:00 VN time
    id='auto_register_meals',
    name='Tự động đăng ký suất ăn 30 ngày',
    replace_existing=True
)
scheduler.start()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5002)



